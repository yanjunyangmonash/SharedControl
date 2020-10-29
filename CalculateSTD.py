import openpyxl
import numpy as np

wb = openpyxl.load_workbook("Clip16_AngleTest5120-5640SL.xlsx")
sheet = wb.active
max_row = sheet.max_row+1
start_row = 0
for x in range(2, max_row):
    if (sheet.cell(row=x, column=1)).value == 'No.2':
        start_row = x
        break
print(start_row)
max_row = sheet.max_row+1

width_list = []
length_list = []
ratio_list = []
for x in range(start_row, max_row):
    width_val = (sheet.cell(row=x, column=2)).value
    if width_val == None:
        continue
    length_val = (sheet.cell(row=x, column=3)).value
    ratio_val = (sheet.cell(row=x, column=4)).value
    width_list.append(width_val)
    length_list.append(length_val)
    ratio_list.append(ratio_val)

width_std = np.std(width_list, ddof=1)
width_mean = np.mean(width_list)
length_std = np.std(length_list, ddof=1)
length_mean = np.mean(length_list)
ratio_std = np.std(ratio_list, ddof=1)
ratio_mean = np.mean(ratio_list)
print(width_mean)
print(width_std)
print(length_mean)
print(length_std)
print(ratio_mean)
print(ratio_std)
#print(str(sheet.max_row))