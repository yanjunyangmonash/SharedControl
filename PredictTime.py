import openpyxl

wb = openpyxl.load_workbook("Clip16_AngleTest5120-5640SL.xlsx")
sheet = wb.active

Max_width = 40.01
Avg_width = 24.44
Min_width = 8.86

Max_length = 48.23
Avg_length = 36.35
Min_length = 24.47

count_out_box_length = 0
count_out_box_width = 0

Time_threshold = 50
zoom_in_frame_a = 0
zoom_in_frame_b = 0
zoom_in_frame_c = 0

max_row = sheet.max_row+1

Found_tool = 0

zoom_in = []

for x in range(2, max_row):
    width_val = (sheet.cell(row=x, column=2)).value
    length_val = (sheet.cell(row=x, column=3)).value

    if Found_tool == 0:
        if width_val * length_val != 0:
            Found_tool = 1

    else:
        if width_val < Min_width or width_val > Max_width:
            count_out_box_width += 1
        else:
            count_out_box_width = 0
        if length_val < Min_length or length_val > Max_length:
            count_out_box_length += 1
        else:
            count_out_box_length = 0
        if count_out_box_length > Time_threshold and count_out_box_width > Time_threshold:
            count_out_box_width = 0
            count_out_box_length = 0
            zoom_in_frame_a = (sheet.cell(row=x, column=1)).value
            zoom_in.append(1)
            print("Zoom in based on case A at frame" + " " + zoom_in_frame_a)
        if count_out_box_length > Time_threshold:
            count_out_box_width = 0
            count_out_box_length = 0
            zoom_in_frame_b = (sheet.cell(row=x, column=1)).value
            zoom_in.append(1)
            print("Zoom in based on case B at frame" + " " + zoom_in_frame_b)
        if count_out_box_width > Time_threshold:
            count_out_box_width = 0
            count_out_box_length = 0
            zoom_in_frame_c = (sheet.cell(row=x, column=1)).value
            zoom_in.append(1)
            print("Zoom in based on case C at frame" + " " + zoom_in_frame_c)

if not len(zoom_in):
    print("No zoom in operation")



