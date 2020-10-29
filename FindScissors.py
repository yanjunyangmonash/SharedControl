import cv2
import openpyxl
from math import atan, tan, asin
import numpy as np
import constant


def find_angle_bisector(line1_p1, line1_p2, line2_p1, line2_p2, centers_matrix):
    # line1 points need to give points on the right
    if (line1_p1[0] - line1_p2[0]) == 0:
        return 0, (0, 0)
    k1 = (line1_p1[1] - line1_p2[1]) / (line1_p1[0] - line1_p2[0])
    a1 = atan(k1)
    b1 = b2 = -1
    c1 = line1_p1[1] - k1 * line1_p1[0]
    # Use perpendicular point to prevent noise (frame627, which choose right point as the next point)
    pvx = (b1 * b1 * line2_p1[0] - k1 * b1 * line2_p1[1] - k1 * c1) / (k1 * k1 + b1 * b1)
    pvy = (-k1 * b1 * line2_p1[0] + k1 * k1 * line2_p1[1] - b1 * c1) / (k1 * k1 + b1 * b1)

    min_line2_p2_dist = 1000
    # If the nearest point pointing to a wrong direction (on the right of the line line2_p1 to pv)
    if (line2_p1[0] - line2_p2[0]) * (pvy - line2_p2[1]) - (line2_p1[1] - line2_p2[1]) * (pvx - line2_p2[0]) > 0:
        wrong_point = line2_p2[0]
        for k in range(len(centers_matrix)):
            new_line2_p2_dist = (([k][0] - line2_p1[0]) ** 2 + (
                    centers_matrix[k][1] - line2_p1[1]) ** 2) ** 0.5
            if new_line2_p2_dist < min_line2_p2_dist and new_line2_p2_dist != 0 and centers_matrix[k][0] != wrong_point:
                new_line2_p2x = centers_matrix[k][0]
                new_line2_p2y = centers_matrix[k][1]
                min_line2_p2_dist = new_line2_p2_dist
        line2_p2 = (new_line2_p2x, new_line2_p2y)

    if (line2_p2[0] - line2_p1[0]) == 0:
        k2 = 0
    else:
        k2 = (line2_p2[1] - line2_p1[1]) / (line2_p2[0] - line2_p1[0])
    a2 = atan(k2)

    if a2 > a1:
        a5 = (a2 - a1) / 2 + a1
    else:
        a5 = (a1 - a2) / 2 + a2
    a1 = k1
    a2 = k2
    c2 = line2_p1[1] - k2 * line2_p1[0]
    k5 = tan(a5)

    if (a1 * b2 - a2 * b1) == 0 or (b1 * a2 - b2 * a1) == 0:
        return k5, (0, 0)

    p5 = ((b1 * c2 - b2 * c1) / (a1 * b2 - a2 * b1), (a1 * c2 - a2 * c1) / (b1 * a2 - b2 * a1))

    return k5, p5


def get_symmetric_point_a(tip_point, edge_point1, edge_point2, k_bisector):
    # Situation a is for bottom; top and curve boundary
    edge_point1_dist = ((edge_point1[1] - tip_point[1]) ** 2 + (edge_point1[0] - tip_point[0]) ** 2) ** 0.5
    edge_point2_dist = ((edge_point2[1] - tip_point[1]) ** 2 + (edge_point2[0] - tip_point[0]) ** 2) ** 0.5

    if edge_point2_dist > edge_point1_dist:
        edge_point1x = ((1 - k_bisector ** 2) * edge_point2[0] - 2 * k_bisector * (-1) * edge_point2[
            1] - 2 * k_bisector * (tip_point[1] - k_bisector * tip_point[0])) / (
                               1 + k_bisector ** 2)
        edge_point1y = ((k_bisector ** 2 - 1) * edge_point2[1] - 2 * k_bisector * (-1) * edge_point2[0] - 2 * (-1) * (
                tip_point[1] - k_bisector * tip_point[0])) / (
                               1 + k_bisector ** 2)
        edge_point1 = (edge_point1x, edge_point1y)
    else:
        edge_point2x = ((1 - k_bisector ** 2) * edge_point1[0] - 2 * k_bisector * (-1) * edge_point1[
            1] - 2 * k_bisector * (tip_point[1] - k_bisector * tip_point[0])) / (
                               1 + k_bisector ** 2)
        edge_point2y = ((k_bisector ** 2 - 1) * edge_point1[1] - 2 * k_bisector * (-1) * edge_point1[0] - 2 * (-1) * (
                tip_point[1] - k_bisector * tip_point[0])) / (
                               1 + k_bisector ** 2)
        edge_point2 = (edge_point2x, edge_point2y)

    return edge_point1, edge_point2


def get_symmetric_point_b(tip_point, edge_point1, k_bisector):
    # Situation b is for bottom&curve and top&curve boundary
    edge_point2x = ((1 - k_bisector ** 2) * edge_point1[0] - 2 * k_bisector * (-1) * edge_point1[1] - 2 * k_bisector * (
                tip_point[1] - k_bisector * tip_point[0])) / (
                           1 + k_bisector ** 2)
    edge_point2y = ((k_bisector ** 2 - 1) * edge_point1[1] - 2 * k_bisector * (-1) * edge_point1[0] - 2 * (-1) * (
                tip_point[1] - k_bisector * tip_point[0])) / (
                           1 + k_bisector ** 2)
    edge_point2 = (edge_point2x, edge_point2y)

    return edge_point2


def get_feature_points_a_test(centers_matrix, circle_rad, circle_x_coor, circle_y_coor):
    list_dist_to_center = []
    for m in range(len(centers_matrix)):
        dist_to_center = circle_rad - abs(
            ((centers_matrix[m][0] - circle_x_coor) ** 2 + (centers_matrix[m][1] - circle_y_coor) ** 2) ** 0.5)
        list_dist_to_center.append(dist_to_center)
    sorted_list = sorted(list_dist_to_center)
    max_number = list_dist_to_center.index(sorted_list[0])
    sec_max_number = list_dist_to_center.index(sorted_list[1])
    max_point = (centers_matrix[max_number][0], centers_matrix[max_number][1])
    next_max_point = get_next_point(centers_matrix, (centers_matrix[max_number][0], centers_matrix[max_number][1]))
    if next_max_point[1] == centers_matrix[sec_max_number][1]:
        sec_max_number = list_dist_to_center.index(sorted_list[2])

    cw = [centers_matrix[max_number][0], centers_matrix[max_number][1]] # max point
    ccw = [centers_matrix[sec_max_number][0], centers_matrix[sec_max_number][1]] # sec max point
    cv2.circle(frame1, (int(cw[0]), int(cw[1])), radius=3, color=(0, 255, 0), thickness=-1)
    cv2.circle(frame1, (int(ccw[0]), int(ccw[1])), radius=3, color=(0, 0, 255), thickness=-1)
    cv2.circle(frame1, (666, 370), radius=3, color=(255, 0, 0), thickness=-1)
    cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
    cwl = []
    ccwl = []

    if centers_matrix[max_number][1] > centers_matrix[sec_max_number][1]:
        cw = [centers_matrix[sec_max_number][0], centers_matrix[sec_max_number][1]]
        ccw = [centers_matrix[max_number][0], centers_matrix[max_number][1]]

    centers_matrix = centers_matrix.reshape([1, -1])
    centers_matrix = centers_matrix.flatten()
    id1 = np.where(centers_matrix == cw[0])
    centers_matrix = np.delete(centers_matrix, id1[0])
    centers_matrix = np.delete(centers_matrix, id1[0])
    id2 = np.where(centers_matrix == ccw[0])
    centers_matrix = np.delete(centers_matrix, id2[0])
    centers_matrix = np.delete(centers_matrix, id2[0])
    centers_matrix = centers_matrix.reshape([-1, 2])
    #centers_matrix = centers_matrix[centers_matrix != cw[0]]
    #centers_matrix = centers_matrix[centers_matrix != cw[1]]
    #centers_matrix = centers_matrix[centers_matrix != ccw[0]]
    #centers_matrix = centers_matrix[centers_matrix != ccw[1]]


    #new_centers = centers_matrix[centers_matrix != cw]
    #new_centers = new_centers.reshape([-1, 2])
    #new_centers = new_centers[new_centers != ccw]
    new_centers = centers_matrix
    new_centers = sorted(new_centers, key=lambda s: s[0], reverse=True)

    for p in new_centers[0:]:
        if ((cw[1]-p[1]) ** 2 + (cw[0]-p[0]) ** 2) ** 0.5 < ((ccw[1]-p[1]) ** 2 + (ccw[0]-p[0]) ** 2) ** 0.5:
            cwl.append(p)
        else:
            ccwl.append(p)
    #if cw[0] == centers_matrix[max_number][0]:
        #next_sec_max_point = (ccwl[0][0], ccwl[0][1])
    #else:
        #next_sec_max_point = (cwl[0][0], cwl[0][1])
    if len(cwl) == 0:
        cwl.append(ccwl[-1])
    if len(ccwl) == 0:
        ccwl.append(cwl[-1])
    cv2.circle(frame1, (int(cwl[0][0]), int(cwl[0][1])), radius=3, color=(0, 255, 0), thickness=-1)
    cv2.circle(frame1, (int(ccwl[0][0]), int(ccwl[0][1])), radius=3, color=(0, 0, 255), thickness=-1)
    cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)

    if cw[1] < ccw[1]:
        right_top_point = (cw[0], cw[1])
        next_right_top_point = (cwl[0][0], cwl[0][1])
        right_bottom_point = (ccw[0], ccw[1])
        next_right_bottom_point = (ccwl[0][0], ccwl[0][1])
    else:
        right_top_point = (ccw[0], ccw[1])
        next_right_top_point = (ccwl[0][0], ccwl[0][1])
        right_bottom_point = (cw[0], cw[1])
        next_right_bottom_point = (cwl[0][0], cwl[0][1])


    return right_top_point, next_right_top_point, right_bottom_point, next_right_bottom_point



def get_feature_points_a(centers_matrix, circle_rad, circle_x_coor, circle_y_coor):
    # Situation a is for curve boundary
    list_dist_to_center = []
    for m in range(len(centers_matrix)):
        dist_to_center = circle_rad - abs(
            ((centers_matrix[m][0] - circle_x_coor) ** 2 + (centers_matrix[m][1] - circle_y_coor) ** 2) ** 0.5)
        list_dist_to_center.append(dist_to_center)
    sorted_list = sorted(list_dist_to_center)
    max_number = list_dist_to_center.index(sorted_list[0])
    sec_max_number = list_dist_to_center.index(sorted_list[1])
    next_max_point = get_next_point(centers_matrix, (centers_matrix[max_number][0], centers_matrix[max_number][1]))
    if next_max_point[1] == centers_matrix[sec_max_number][1]:
        sec_max_number = list_dist_to_center.index(sorted_list[2])
    next_sec_max_point = get_next_point(centers_matrix, (centers_matrix[sec_max_number][0], centers_matrix[sec_max_number][1]), restriction_point=(centers_matrix[max_number][0], centers_matrix[max_number][1]))
    #cv2.circle(frame1, (int(centers_matrix[max_number][0]), int(centers_matrix[max_number][1])), radius=3,
               #color=(255, 255, 0), thickness=-1)
    #cv2.circle(frame1, (int(next_max_point[0]), int(next_max_point[1])), radius=3, color=(255, 255, 0),
               #thickness=-1)
    #cv2.circle(frame1, (int(centers_matrix[sec_max_number][0]), int(centers_matrix[sec_max_number][1])), radius=3, color=(255, 0, 0), thickness=-1)
    #cv2.circle(frame1, (int(next_sec_max_point[0]), int(next_sec_max_point[1])), radius=3, color=(255, 0, 0), thickness=-1)
    #cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)

    if centers_matrix[max_number][1] < centers_matrix[sec_max_number][1]:
        right_top_point = (centers_matrix[max_number][0], centers_matrix[max_number][1])
        next_right_top_point = next_max_point
        right_bottom_point = (centers_matrix[sec_max_number][0], centers_matrix[sec_max_number][1])
        next_right_bottom_point = next_sec_max_point
    else:
        right_top_point = (centers_matrix[sec_max_number][0], centers_matrix[sec_max_number][1])
        next_right_top_point = next_sec_max_point
        right_bottom_point = (centers_matrix[max_number][0], centers_matrix[max_number][1])
        next_right_bottom_point = next_max_point

    return right_top_point, next_right_top_point, right_bottom_point, next_right_bottom_point

def get_feature_points_b(centers_matrix, bottom_or_top=1):
    # Situation b is for bottom&curve and top&curve boundary
    list_points_x = sorted(centers_matrix, key=lambda s: s[0])
    list_points_y = sorted(centers_matrix, key=lambda s: s[1])
    right_point = (list_points_x[-1][0], list_points_x[-1][1])
    if bottom_or_top:
        # Find bottom point
        top_or_bottom_point = (list_points_y[-1][0], list_points_y[-1][1])
        if top_or_bottom_point[1] <= right_point[1]:
            top_or_bottom_point = (list_points_y[-2][0], list_points_y[-2][1])
    else:
        # Find top point
        top_or_bottom_point = (list_points_y[0][0], list_points_y[0][1])
        if top_or_bottom_point[1] >= right_point[1]:
            top_or_bottom_point = (list_points_y[1][0], list_points_y[1][1])

    return right_point, top_or_bottom_point


def get_feature_points_c(centers_matrix, right_boundary, left_boundary, bottom_or_top=1):
    # Situation c is for bottom and top boundary
    if bottom_or_top:
        list_points_y = sorted(centers_matrix, key=lambda s: s[1], reverse=True)
        bottom_points = [list_points_y[0], list_points_y[1], list_points_y[2]]
        list_bottom_points_x = sorted(bottom_points, key=lambda s: s[0], reverse=True)
        right_bottom_point = (list_bottom_points_x[0][0], list_bottom_points_x[0][1])
        next_right_bottom_point = get_next_point(centers_matrix, right_bottom_point, add_x_restriction=1)
        if bottom_points[1][1] != next_right_bottom_point[1]:
            bottom_point = (bottom_points[1][0], bottom_points[1][1])
        else:
            bottom_point = (bottom_points[2][0], bottom_points[2][1])
        next_bottom_point = get_next_point(centers_matrix, bottom_point, add_x_restriction=0)

        '''
        # For bottom situation
        list_points_y = sorted(centers_matrix, key=lambda s: s[1], reverse=True)
        bottom_point = (list_points_y[0][0], list_points_y[0][1])
        min_dist = 1000
        for m in range(len(centers_matrix)):
            dist_to_right_bottom = ((centers_matrix[m][0] - right_boundary) ** 2 + (centers_matrix[m][1] - 540) ** 2) ** 0.5
            if dist_to_right_bottom < min_dist:
                right_bottom_point = (centers_matrix[m][0], centers_matrix[m][1])
                min_dist = dist_to_right_bottom
        next_right_bottom_point = get_next_point(centers_matrix, right_bottom_point, add_x_restriction=1)

        if bottom_point[0] == right_bottom_point[0] or bottom_point[0] == next_right_bottom_point[0]:
            for i in range(len(list_points_y)):
                if list_points_y[i][1] != right_bottom_point[1] and list_points_y[i][1] != next_right_bottom_point[1]:
                    bottom_point = (list_points_y[i][0], list_points_y[i][1])
                    break
        next_bottom_point = get_next_point(centers_matrix, bottom_point, add_x_restriction=0)
        '''
        output_point1 = right_bottom_point
        output_point2 = next_right_bottom_point
        output_point3 = bottom_point
        output_point4 = next_bottom_point

    else:
        list_points_y = sorted(centers_matrix, key=lambda s: s[1])
        cv2.circle(frame1, (int(list_points_y[0][0]), int(list_points_y[0][1])), 8, (255, 0, 0), 5)
        cv2.circle(frame1, (int(list_points_y[1][0]), int(list_points_y[1][1])), 8, (0, 255, 0), 5)
        cv2.circle(frame1, (int(list_points_y[2][0]), int(list_points_y[2][1])), 8, (0, 0, 255), 5)
        cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
        top_points = [list_points_y[0], list_points_y[1], list_points_y[2]]
        list_top_points_x = sorted(top_points, key=lambda s: s[0], reverse=True)

        # Test -------------------------
        mid_to_right = list_top_points_x[0][0] - list_top_points_x[1][0]
        mid_to_left = list_top_points_x[2][0] - list_top_points_x[1][0]

        if mid_to_right < mid_to_left:
            if list_top_points_x[0][1] < list_top_points_x[1][1]:
                right_top_point = (list_top_points_x[0][0], list_top_points_x[0][1])
                del list_top_points_x[0]
            else:
                right_top_point = (list_top_points_x[1][0], list_top_points_x[1][1])
                del list_top_points_x[1]
        else:
            right_top_point = (list_top_points_x[0][0], list_top_points_x[0][1])
            del list_top_points_x[0]
        # Test -------------------------

        #right_top_point = (list_top_points_x[0][0], list_top_points_x[0][1])
        next_right_top_point = get_next_point(centers_matrix, right_top_point, add_x_restriction=1)

        # Test -------------------------
        dist1 = (next_right_top_point[0]-list_top_points_x[0][0])**2 + (next_right_top_point[1]-list_top_points_x[0][1])**2
        dist2 = (next_right_top_point[0]-list_top_points_x[1][0])**2 + (next_right_top_point[1]-list_top_points_x[1][1])**2
        if dist1 < dist2:
            del list_top_points_x[0]
        else:
            del list_top_points_x[1]

        #if top_points[1][1] != next_right_top_point[1]:
            #top_point = (top_points[1][0], top_points[1][1])
        #else:
            #top_point = (top_points[2][0], top_points[2][1])
        top_point = (list_top_points_x[0][0], list_top_points_x[0][1])
        next_top_point = get_next_point(centers_matrix, top_point, add_x_restriction=0)

        '''
        list_points_y = sorted(centers_matrix, key=lambda s: s[1])
        top_point = (list_points_y[0][0], list_points_y[0][1])
        min_dist = 1000
        for m in range(len(centers_matrix)):
            dist_to_left_top = ((centers_matrix[m][0] - left_boundary) ** 2 + (
                        centers_matrix[m][1] - 0) ** 2) ** 0.5
            if dist_to_left_top < min_dist:
                left_top_point = (centers_matrix[m][0], centers_matrix[m][1])
                min_dist = dist_to_left_top
        next_left_top_point = get_next_point(centers_matrix, left_top_point, add_x_restriction=1)

        if top_point[0] == left_top_point[0] or top_point[0] == next_left_top_point[0]:
            for i in range(len(list_points_y)):
                if list_points_y[i][1] != left_top_point[1] and list_points_y[i][1] != next_left_top_point[1]:
                    top_point = (list_points_y[i][0], list_points_y[i][1])
                    break
        next_top_point = get_next_point(centers_matrix, top_point, add_x_restriction=0)
        
        output_point1 = left_top_point
        output_point2 = next_left_top_point
        output_point3 = top_point
        output_point4 = next_top_point
        '''
        output_point1 = top_point
        output_point2 = next_top_point
        output_point3 = right_top_point
        output_point4 = next_right_top_point

    return output_point1, output_point2, output_point3, output_point4



def get_next_point(centers_matrix, end_point, add_x_restriction=1, restriction_point = (0,0)):
    next_point = (0, 0)
    min_dist = 1000
    if add_x_restriction:
        if restriction_point == (0, 0):
            for k in range(len(centers_matrix)):
                dist_to_next_point = ((centers_matrix[k][0] - end_point[0]) ** 2 + (centers_matrix[k][1] - end_point[1]) ** 2) ** 0.5
                if dist_to_next_point < min_dist and dist_to_next_point != 0 and abs((end_point[1] - centers_matrix[k][1])) < 130 and centers_matrix[k][0] < end_point[0]:
                    next_point = (centers_matrix[k][0], centers_matrix[k][1])
                    min_dist = dist_to_next_point
        else:
            for k in range(len(centers_matrix)):
                dist_to_next_point = ((centers_matrix[k][0] - end_point[0]) ** 2 + (centers_matrix[k][1] - end_point[1]) ** 2) ** 0.5
                if dist_to_next_point < min_dist and dist_to_next_point != 0 and abs(
                    (end_point[1] - centers_matrix[k][1])) < 130 and centers_matrix[k][0] < end_point[0] and centers_matrix[k][1] != restriction_point[1]:
                    next_point = (centers_matrix[k][0], centers_matrix[k][1])
                    min_dist = dist_to_next_point
    else:
        if restriction_point == (0, 0):
            for k in range(len(centers_matrix)):
                dist_to_next_point = ((centers_matrix[k][0] - end_point[0]) ** 2 + (centers_matrix[k][1] - end_point[1]) ** 2) ** 0.5
                if dist_to_next_point < min_dist and dist_to_next_point != 0 and abs((end_point[1] - centers_matrix[k][1])) < 130:
                    next_point = (centers_matrix[k][0], centers_matrix[k][1])
                    min_dist = dist_to_next_point
        else:
            for k in range(len(centers_matrix)):
                dist_to_next_point = ((centers_matrix[k][0] - end_point[0]) ** 2 + (centers_matrix[k][1] - end_point[1]) ** 2) ** 0.5
                if dist_to_next_point < min_dist and dist_to_next_point != 0 and abs(
                    (end_point[1] - centers_matrix[k][1])) < 130 and centers_matrix[k][1] != restriction_point[1]:
                    next_point = (centers_matrix[k][0], centers_matrix[k][1])
                    min_dist = dist_to_next_point

    return next_point


def prep_for_Kmeans(contours_number, max_id, circle_x_coor, circle_y_coor, left_bound, right_bound, i_r, o_r, height):
    curve_pathx = []
    curve_pathy = []
    upbox_x = []
    upbox_y = []
    downbox_x = []
    downbox_y = []
    points_group = []

    for j in range(len(contours_number[max_id])):
        dist = (contours_number[max_id][j][0][0] - circle_x_coor) ** 2 + (
                contours_number[max_id][j][0][1] - circle_y_coor) ** 2
        if dist <= i_r ** 2:
            if height + 10 < contours_number[max_id][j][0][1] < 540 - (height + 10):
                points_group.append(contours_number[max_id][j])
        if i_r ** 2 <= dist <= o_r ** 2:
            curve_pathx.append(contours_number[max_id][j][0][0])
            curve_pathy.append(contours_number[max_id][j][0][1])
        if (left_bound + 15 < contours_number[max_id][j][0][0] < right_bound - 15) and (
                0 <= contours_number[max_id][j][0][1] <= height):
            upbox_x.append(contours_number[max_id][j][0][0])
            upbox_y.append(contours_number[max_id][j][0][1])
        if (left_bound + 15 < contours[max_id][j][0][0] < right_bound - 15) and (
                540 - height <= contours[max_id][j][0][1] <= 540):
            downbox_x.append(contours_number[max_id][j][0][0])
            downbox_y.append(contours_number[max_id][j][0][1])

    return points_group, curve_pathx, curve_pathy, upbox_x, upbox_y, downbox_x, downbox_y


def kmeans_algorithm(points_group):
    points_group = np.float32(points_group)
    points_group = np.array(points_group)
    clusterCount = 12  # Number of clusters to split the set by
    attempts = 10  # Number of times the algorithm is executed using different initial labels
    flags = cv2.KMEANS_PP_CENTERS
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1.0)
    if len(points_group) < clusterCount:
        return [], 0
    compactness, labels, centers = cv2.kmeans(points_group, clusterCount, None, criteria, attempts, flags)

    return centers, 1


def get_tool_length(point2, point3, centers_matrix, k_val):
    if point2[0] > point3[0]:
        mid_point23 = (int(point3[0] + (point2[0] - point3[0]) / 2), int(point3[1] + (point2[1] - point3[1]) / 2))
    else:
        mid_point23 = (int(point2[0] + (point3[0] - point2[0]) / 2), int(point2[1] + (point3[1] - point2[1]) / 2))
    dist_list = []
    for point_id in range(len(centers_matrix)):
        dist = ((centers_matrix[point_id][0] - mid_point23[0]) ** 2 + (centers_matrix[point_id][1] - mid_point23[1]) ** 2) ** 0.5
        dist_list.append(dist)
    tip_id = dist_list.index(max(dist_list))
    tip_point = (centers_matrix[tip_id][0], centers_matrix[tip_id][1])

    a1 = k_val
    b1 = -1
    c1 = tip_point[1] - k_val * tip_point[0]
    if point2[0]-point3[0] == 0:
        return tip_point, (0, 0), 0
    a2 = (point2[1]-point3[1])/(point2[0]-point3[0])
    b2 = -1
    c2 = point2[1] - a2 * point2[0]

    if (a1 * b2 - a2 * b1) == 0 or (b1 * a2 - b2 * a1) == 0:
        return tip_point, (0, 0), 0

    intersect_point = ((b1 * c2 - b2 * c1) / (a1 * b2 - a2 * b1), (a1 * c2 - a2 * c1) / (b1 * a2 - b2 * a1))
    length = ((tip_point[0] - intersect_point[0]) ** 2 + (tip_point[1] - intersect_point[1]) ** 2) ** 0.5

    return tip_point, intersect_point, length



def calculate_distances(no_of_frame, contours_number, circle_x_coor, circle_y_coor, left_bound, right_bound, excel_number, inner_r, outer_r, real_r, box_height, pre_width, pre_length, pre_LW_Ratio):
    # Set up parameters
    contour_areas = []
    two_tools_touch = 0
    true_mass_xs = []
    true_mass_ys = []
    number = []
    RowNumber = excel_number
    ColumnNumber = 1

    for num_of_contours in range(len(contours_number)):
        M = cv2.moments(contours_number[num_of_contours], 0)
        if M['m00']:
            mass_centres_x = int(M['m10'] / M['m00'])
            mass_centres_y = int(M['m01'] / M['m00'])
        else:
            continue

        if mass_centres_x < circle_x_coor:
            number.append(0)
            contour_areas.append(0)
            true_mass_xs.append(0)
            true_mass_ys.append(0)
            numbers = len(contours_number[num_of_contours])
            for i in range(numbers):
                if contours_number[num_of_contours][i][0][0] > boundaryr:
                    dist = (contours_number[num_of_contours][i][0][0] - circle_x) ** 2 + (
                            contours_number[num_of_contours][i][0][1] - circle_y) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_rad ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(no_of_frame)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        two_tools_touch = 1
                        break

        elif mass_centres_x >= circle_x_coor:
            # Calculate min area rect
            rect = cv2.minAreaRect(contours_number[num_of_contours])
            box = cv2.boxPoints(rect)
            box_h = abs(((box[0][0] - box[1][0]) ** 2 + (box[0][1] - box[1][1])) ** 0.5)
            box_w = abs(((box[1][0] - box[2][0]) ** 2 + (box[1][1] - box[2][1])) ** 0.5)
            number.append(num_of_contours)
            contour_areas.append(box_h*box_w)
            true_mass_xs.append(mass_centres_x)
            true_mass_ys.append(mass_centres_y)

            numbers = len(contours_number[num_of_contours])
            for i in range(numbers):
                if contours_number[num_of_contours][i][0][0] < left_bound:
                    dist = (contours_number[num_of_contours][i][0][0] - circle_x_coor) ** 2 + (
                            contours_number[num_of_contours][i][0][1] - circle_y_coor) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_r ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        two_tools_touch = 1
                        break


    if two_tools_touch == 1:
        two_tools_touch = 0
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    if len(contour_areas) == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    if max(contour_areas) < 300 or len(contour_areas) == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    # Find max contour on the right side
    max_num_id = contour_areas.index(max(contour_areas))
    true_mass_x = true_mass_xs[max_num_id]
    true_mass_y = true_mass_ys[max_num_id]
    max_num = number[max_num_id]

    # Collect points from the max contour to prepare K-means
    points, curve_pathx, curve_pathy, upbox_x, upbox_y, downbox_x, downbox_y = prep_for_Kmeans(contours_number, max_num, circle_x_coor, circle_y_coor, left_bound,
                    right_bound, inner_r, outer_r, box_height)
    # Run the K means to get feature points
    centers, have_centers = kmeans_algorithm(points)
    if have_centers == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio
    for l in range(len(centers)):
        cv2.circle(frame1, (int(centers[l][0]), int(centers[l][1])), radius=3, color=(0, 255, 0), thickness=-1)

    #cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)

    # Locating the mask by using three points (two edges and one tip)
    p4 = (true_mass_x, true_mass_y)

    if len(downbox_x) and len(curve_pathx):
        pr, pd = get_feature_points_b(centers, bottom_or_top=1)
        pnr = get_next_point(centers, pr, add_x_restriction=1)
        pnd = get_next_point(centers, pd, add_x_restriction=0)
        k5, p5 = find_angle_bisector(pr, pnr, pd, pnd, centers)
        p3_index = curve_pathy.index(min(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2 = get_symmetric_point_b(p5, p3, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pd[0]), int(pd[1])), (int(pnd[0]), int(pnd[1])), (255, 0, 0), thickness=5)

    elif len(downbox_x) and len(curve_pathx) == 0:
        pr, pnr, pl, pnl = get_feature_points_c(centers, right_bound, left_bound, bottom_or_top=1)
        k5, p5 = find_angle_bisector(pr, pnr, pl, pnl, centers)
        p2_index = downbox_x.index(min(downbox_x))
        p2 = (downbox_x[p2_index], downbox_y[p2_index])
        p3_index = downbox_x.index(max(downbox_x))
        p3 = (downbox_x[p3_index], downbox_y[p3_index])
        p2, p3 = get_symmetric_point_a(p5, p3, p2, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pl[0]), int(pl[1])), (int(pnl[0]), int(pnl[1])), (255, 0, 0), thickness=5)

    elif len(upbox_x) and len(curve_pathx):
        pr, pu = get_feature_points_b(centers, bottom_or_top=0)
        pnr = get_next_point(centers, pr, add_x_restriction=1)
        pnu = get_next_point(centers, pu, add_x_restriction=0)
        k5, p5 = find_angle_bisector(pu, pnu, pr, pnr, centers)
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2 = get_symmetric_point_b(p5, p3, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pu[0]), int(pu[1])), (int(pnu[0]), int(pnu[1])), (255, 0, 0), thickness=5)

    elif len(upbox_x) and len(curve_pathx) == 0:
        pl, pnl, pr, pnr = get_feature_points_c(centers, right_bound, left_bound, bottom_or_top=0)
        k5, p5 = find_angle_bisector(pl, pnl, pr, pnr, centers)
        p2_index = upbox_x.index(min(upbox_x))
        p2 = (upbox_x[p2_index], upbox_y[p2_index])
        p3_index = upbox_x.index(max(upbox_x))
        p3 = (upbox_x[p3_index], upbox_y[p3_index])
        p2, p3 = get_symmetric_point_a(p5, p3, p2, k5)

        #cv2.line(frame1, (int(pl[0]), int(pl[1])), (int(pnl[0]), int(pnl[1])), (255, 255, 0), thickness=5)
        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 255), thickness=5)
        cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)

    elif len(curve_pathx):
        pu, pnu, pd, pnd = get_feature_points_a(centers, real_r, circle_x_coor, circle_y_coor)
        k5, p5 = find_angle_bisector(pu, pnu, pd, pnd, centers)
        p2_index = curve_pathy.index(min(curve_pathy))
        p2 = (curve_pathx[p2_index], curve_pathy[p2_index])
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2, p3 = get_symmetric_point_a(p5, p3, p2, k5)

        cv2.line(frame1, (int(pd[0]), int(pd[1])), (int(pnd[0]), int(pnd[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pu[0]), int(pu[1])), (int(pnu[0]), int(pnu[1])), (0, 0, 255), thickness=5)

    else:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio


    pointsdist = ((p2[1] - p3[1]) ** 2 + (p2[0] - p3[0]) ** 2) ** 0.5
    arclength = asin(circle_y/real_r)*2*real_r
    pointsdist = pointsdist/arclength*100
    pre_width = pointsdist

    p_tip1, p_tip2, length = get_tool_length(p2, p3, centers, k5)
    length = length / (real_r * 2) * 100
    pre_length = length
    if pointsdist != 0:
        LW_Ratio = length/pointsdist
        if LW_Ratio > 20:
            LW_Ratio = 0
    else:
        LW_Ratio = 0
    pre_LW_Ratio = LW_Ratio

    sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
    sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pointsdist)
    sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=length)
    sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=LW_Ratio)
    RowNumber += 1

    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p2[0]), int(p2[1])), (0, 0, 255), thickness=1)
    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p3[0]), int(p3[1])), (0, 0, 255), thickness=1)
    cv2.line(frame1, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), thickness=2)
    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p5[0]) + 900, int(p5[1] + 900 * k5)), (255, 0, 0), thickness=2)
    cv2.line(frame1, (int(p_tip1[0]), int(p_tip1[1])), (int(p_tip2[0]), int(p_tip2[1])), (0, 0, 255), thickness=2)

    #for l in range(len(centers)):
        #cv2.circle(frame1, (int(centers[l][0]), int(centers[l][1])), radius=3, color=(0, 255, 0), thickness=-1)

    cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
    print('No.' + str(frames))
    cv2.waitKey(0)

    return RowNumber, pre_width, pre_length, pre_LW_Ratio

if __name__ == "__main__":
    # Excel setup
    tablepath = 'Clip16_AngleTest5120-5640SL.xlsx'
    Row_Number = 2
    ColumnNumber = 1
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=ColumnNumber, value='Img')
    sheet.cell(row=1, column=ColumnNumber + 1, value='Width')
    sheet.cell(row=1, column=ColumnNumber + 2, value='Length')
    sheet.cell(row=1, column=ColumnNumber + 3, value='LW_Ratio')
    prewidth = 0
    prelength = 0
    prelwratio = 0
    # ---------------------------------

    # Laparoscopic view geo parameters
    video_num = 42
    video_constant = constant.VideoConstants()
    constant_values = video_constant.num_to_constants(video_num)()

    true_rad = constant_values[0]
    inner_rad = constant_values[1]
    outer_rad = constant_values[2]
    circle_x = constant_values[3]
    circle_y = constant_values[4]
    boundaryl = constant_values[5]
    boundaryr = constant_values[6]
    boxheight = 10
    # --------------------------------

    for frames in range(5938, 4945, 1):
        if video_num % 10 == 0:
            folder_name = str(10 * (video_num // 10) - 9) + '-' + str(10 * (video_num // 10))
        else:
            folder_name = str(10*(video_num//10)+1) + '-' + str(10*(video_num//10)+10)
        #print('E:/Clip' + folder_name + '/Clip' + str(video_num) + '_1M/clip' + str(video_num) + '' + str(frames) + 'M.jpg')
        frame = cv2.imread('E:/Clip' + folder_name + '/Clip' + str(video_num) + '_1M/clip' + str(video_num) + '_' + str(frames) + 'M.jpg')
        frame1 = cv2.imread('E:/Clip' + folder_name + '/Clip' + str(video_num) + '_1D/clip' + str(video_num) + '_' + str(frames) + 'D.jpg')
        # Have problem in clip16 frame 5520
        #grayImage = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        #_, bw = cv2.threshold(grayImage, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        #testEdges = cv2.Canny(bw, 30, 30 * 3)
        #contours, _ = cv2.findContours(testEdges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        mask_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        ret, contours_frame = cv2.threshold(mask_frame, 127, 255, cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(contours_frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        Row_Number, prewidth, prelength, prelwratio = calculate_distances(frames, contours, circle_x, circle_y, boundaryl, boundaryr,
                            Row_Number, inner_rad, outer_rad, true_rad, boxheight, prewidth, prelength, prelwratio)

    workbook.save(tablepath)
    print('Finish')

