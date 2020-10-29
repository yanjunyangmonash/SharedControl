import numpy as np
from sklearn.decomposition import PCA
import openpyxl
import cv2
from math import atan2, atan

def getOrientation(pts):
    sz = len(pts)
    data_pts = np.empty((sz, 2), dtype=np.float64)
    for i in range(data_pts.shape[0]):
        data_pts[i, 0] = pts[i, 0, 0]
        data_pts[i, 1] = pts[i, 0, 1]
    # Perform PCA analysis
    mean = np.empty((0))
    mean, eigenvectors, eigenvalues = cv2.PCACompute2(data_pts, mean)
    angle = atan2(eigenvectors[0, 1], eigenvectors[0, 0])  # orientation in radians

    return angle
'''
wb = openpyxl.load_workbook('Clip32_AngleTest1115-2322.xlsx', data_only=True)
active_sheet = wb.active
TArray = []
for i in range(2, 1210):
    cellname1 = 'L' + str(i)
    cellname2 = 'P' + str(i)
    cellname3 = 'S' + str(i)
    cell1 = active_sheet[cellname1]
    cell2 = active_sheet[cellname2]
    cell3 = active_sheet[cellname2]
    TArray.append([cell1.value, cell2.value])
X = np.array(TArray)
pca = PCA(n_components=2)
newX = pca.fit_transform(X)
print(pca.explained_variance_ratio_)

wb = openpyxl.load_workbook('Clip32_AngleTest1115-2322.xlsx', data_only=True)
tablepath = 'Clip32_AngleTest1115-2322PCA.xlsx'
active_sheet = wb.active
TArray = []
count = 0
groups = 0

RowNumber = 2
ColumnNumber = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=ColumnNumber + 1, value='PC1')
sheet.cell(row=1, column=ColumnNumber + 2, value='PC2')

for i in range(2, 1210):
    cellname1 = 'L' + str(i)
    cellname2 = 'P' + str(i)
    cellname3 = 'S' + str(i)
    cell1 = active_sheet[cellname1]
    cell2 = active_sheet[cellname2]
    cell3 = active_sheet[cellname2]
    TArray.append([cell1.value, cell2.value, cell3.value])
    count += 1


    if count == 50:
        X = np.array(TArray)
        pca = PCA(n_components=2)
        newX = pca.fit_transform(X)
        #print(pca.explained_variance_ratio_)
        for j in range(len(newX)):
            sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=newX[j][0])
            sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=newX[j][1])
            RowNumber += 1
        count = 0
        TArray = []
        groups += 1
        print(groups)

    elif groups == (1210 - 2) // 50 and count == (1210 - 2) % 50:
        X = np.array(TArray)
        pca = PCA(n_components=2)
        newX = pca.fit_transform(X)
        for k in range(len(newX)):
            sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=newX[k][0])
            sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=newX[k][1])
            RowNumber += 1
        print('Rest saved')
        print(len(newX))

workbook.save(tablepath)
print('Finish')
'''


frame = cv2.imread('Clip24_1M/clip24' + '_' + str(115) + 'M.jpg')
frame1 = cv2.imread('Clip24_1D/clip24' + '_' + str(115) + 'D.jpg')
frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#ret, frame = cv2.threshold(frame, 127, 255, cv2.THRESH_BINARY)
#contours, hierarchy = cv2.findContours(frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
_, bw = cv2.threshold(frame, 50, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
contours, _ = cv2.findContours(bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
TestArray = []

for num_of_contours in range(len(contours)):
    M = cv2.moments(contours[num_of_contours], 0)
    if M['m00']:
        mass_centres_x = int(M['m10'] / M['m00'])
        mass_centres_y = int(M['m01'] / M['m00'])
    else:
        continue

    if mass_centres_x >= 494:
        numbers = len(contours[num_of_contours])
        for i in range(numbers):
            TestArray.append([[contours[num_of_contours][i][0][0], contours[num_of_contours][i][0][1]]])
TestArray = np.array(TestArray)
angle = getOrientation(TestArray)
#print(angle)
k = atan(angle)
print(k)
#cv2.line(frame1, (594, 441), (594+120, int(441+120*k)), (0, 0, 255), thickness=3)
#cv2.line(frame1, (753, 457), (594, 441), (0, 0, 255), thickness=3)
#cv2.line(frame1, (728, 528), (594, 441), (0, 0, 255), thickness=3)
#cv2.imshow('output', frame1)
#cv2.waitKey()
'''
X = np.array(TestArray)
pca = PCA(n_components=2)
newX = pca.fit_transform(X)
print(pca.explained_variance_ratio_)
cov = pca.get_covariance()
print(cov)
evals, evecs = np.linalg.eig(cov)
print(evals)
sort_indices = np.argsort(evals)[::-1]
x_v1, y_v1 = evecs[:, sort_indices[0]]
print(x_v1)
print(y_v1)
print(evecs)
cv2.line(frame1, (0, 0), (50, -50*int(y_v1/x_v1)), (0, 0, 255), thickness=1)
cv2.imshow('A', frame1)
cv2.waitKey(0)
'''