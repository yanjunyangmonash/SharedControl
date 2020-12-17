import openpyxl


def write_excel_title():
    excelbook = openpyxl.Workbook()
    sheet = excelbook.active
    ColumnNumber = 1
    sheet.cell(row=1, column=ColumnNumber, value='Img')
    sheet.cell(row=1, column=ColumnNumber + 1, value='Width')
    sheet.cell(row=1, column=ColumnNumber + 2, value='Length')
    sheet.cell(row=1, column=ColumnNumber + 3, value='LW_Ratio')
    return excelbook


def save_excel_table(tablepath, excelbook):
    excelbook.save(tablepath)


def write_excel_table(frame, excelbook, row_num, val1=None, val2=None, val3=None):
    sheet = excelbook.active
    columnNumber = 1
    sheet.cell(row=row_num, column=columnNumber, value=('No.' + str(frame)))
    sheet.cell(row=row_num, column=columnNumber + 1, value=val1)
    sheet.cell(row=row_num, column=columnNumber + 2, value=val2)
    sheet.cell(row=row_num, column=columnNumber + 3, value=val3)
