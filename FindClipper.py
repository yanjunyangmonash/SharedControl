import cv2
import openpyxl
from math import atan2, asin, pi
import numpy as np
import constant
from utils import GeoCalculation as GC
from utils import kmeans as k
from utils import symmetricpoints as sp
from utils import featurepoints as fp
from utils import anglebisector as ab
from utils import write_excel as we
from utils import toolselection as ts


def calculate_distances_test(contours_number, excel_number, pre_width, pre_length, pre_LW_Ratio, main_tool,
                             main_tool_coor,
                             assist_tool_coor):
    # Set up parameters
    RowNumber = excel_number
    ColumnNumber = 1

    angle = atan2(270, (boundaryr - circle_x))
    angle = int(angle * 180 / pi)
    lapview_area = 270 * (boundaryr - boundaryl) + (pi * (true_rad ** 2)) * (angle / 180)
    small_area_metrics = lapview_area * 0.03

    two_tools_touch, contour_areas, true_mass_xs, true_mass_ys, number = ts.valid_mask_collector(contours_number)

    if two_tools_touch == 1:
        ts.touched_tools_filter(RowNumber, frames, pre_width, pre_length, pre_LW_Ratio, workbook, frame1)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor

    if len(contour_areas) == 0 or max(contour_areas) < 2:
        ts.left_tool_filter(RowNumber, frames, workbook, frame1)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor

    have_ee, max_num_id, sec_max_num_id = ts.end_effector_filter(frames, frame1, contour_areas, true_mass_xs,
                                                                 true_mass_ys, workbook, RowNumber)
    if have_ee:
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor

    if max(contour_areas) < small_area_metrics:
        ts.noise_filter(frames, frame1, workbook, RowNumber)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor

    # Set the first main tool coordinates
    continue_process, main_tool_coor, assist_tool_coor, max_num_id = ts.main_tool_tracker(RowNumber, frames,
                                                                                          pre_width, workbook,
                                                                                          frame1, main_tool,
                                                                                          contour_areas,
                                                                                          true_mass_xs,
                                                                                          true_mass_ys,
                                                                                          max_num_id,
                                                                                          sec_max_num_id,
                                                                                          main_tool_coor,
                                                                                          assist_tool_coor)
    if continue_process == 0:
        RowNumber += 1
        return RowNumber, None, None, None, main_tool, main_tool_coor, assist_tool_coor

    interrupt_loop = ts.bad_main_tool_mask_filter(contours_number, max_num_id, frames, frame1, workbook, RowNumber)
    if interrupt_loop:
        RowNumber += 1
        return RowNumber, None, None, None, main_tool, main_tool_coor, assist_tool_coor

    # Find max contour on the right side
    true_mass_x = true_mass_xs[max_num_id]
    true_mass_y = true_mass_ys[max_num_id]
    max_num = number[max_num_id]

    # Collect points from the max contour to prepare K-means
    points, curve_pathx, curve_pathy, upbox_x, upbox_y, downbox_x, downbox_y = k.prep_for_Kmeans(contours_number,
                                                                                                 max_num,
                                                                                                 circle_x,
                                                                                                 circle_y,
                                                                                                 boundaryl,
                                                                                                 boundaryr, inner_rad,
                                                                                                 outer_rad, boxheight)
    # Run the K means to get feature points
    centers, have_centers = k.kmeans_algorithm(points)
    if have_centers == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor
    for l in range(len(centers)):
        cv2.circle(frame1, (int(centers[l][0]), int(centers[l][1])), radius=3, color=(0, 255, 0), thickness=-1)

    # Locating the mask by using three points (two edges and one tip)
    p4 = (true_mass_x, true_mass_y)

    if len(downbox_x) and len(curve_pathx):
        pr, pnr, pd, pnd = fp.get_feature_points_b_use_k(centers, bottom_or_top=1)
        k5, p5 = ab.find_angle_bisector_curveedge(pr, pnr, pd, pnd, centers)
        p3_index = curve_pathy.index(min(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2 = sp.get_symmetric_point_b(p5, p3, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pd[0]), int(pd[1])), (int(pnd[0]), int(pnd[1])), (255, 0, 0), thickness=5)

    elif len(downbox_x) and len(curve_pathx) == 0:
        pr, pnr, pl, pnl = fp.get_feature_points_c_use_k(centers, bottom_or_top=1)
        k5, p5 = ab.find_angle_bisector(pr, pnr, pl, pnl, centers)
        p2_index = downbox_x.index(min(downbox_x))
        p2 = (downbox_x[p2_index], downbox_y[p2_index])
        p3_index = downbox_x.index(max(downbox_x))
        p3 = (downbox_x[p3_index], downbox_y[p3_index])
        p2, p3 = sp.get_symmetric_point_a(p5, p3, p2, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pl[0]), int(pl[1])), (int(pnl[0]), int(pnl[1])), (255, 0, 0), thickness=5)

    elif len(upbox_x) and len(curve_pathx):
        pr, pnr, pu, pnu = fp.get_feature_points_b_use_k(centers, bottom_or_top=0)
        k5, p5 = ab.find_angle_bisector_curveedge(pu, pnu, pr, pnr, centers)
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2 = sp.get_symmetric_point_b(p5, p3, k5)

        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pu[0]), int(pu[1])), (int(pnu[0]), int(pnu[1])), (255, 0, 0), thickness=5)

    elif len(upbox_x) and len(curve_pathx) == 0:
        pl, pnl, pr, pnr = fp.get_feature_points_c_use_k(centers, bottom_or_top=0)
        k5, p5 = ab.find_angle_bisector(pl, pnl, pr, pnr, centers)
        p2_index = upbox_x.index(min(upbox_x))
        p2 = (upbox_x[p2_index], upbox_y[p2_index])
        p3_index = upbox_x.index(max(upbox_x))
        p3 = (upbox_x[p3_index], upbox_y[p3_index])
        p2, p3 = sp.get_symmetric_point_a(p5, p3, p2, k5)

        cv2.line(frame1, (int(pl[0]), int(pl[1])), (int(pnl[0]), int(pnl[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pr[0]), int(pr[1])), (int(pnr[0]), int(pnr[1])), (255, 0, 0), thickness=5)

    elif len(curve_pathx):
        pu, pnu, pd, pnd = fp.get_feature_points_a_use_k(centers, true_rad, circle_x, circle_y)
        k5, p5 = ab.find_angle_bisector_curve(pu, pnu, pd, pnd, centers)
        p2_index = curve_pathy.index(min(curve_pathy))
        p2 = (curve_pathx[p2_index], curve_pathy[p2_index])
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2, p3 = sp.get_symmetric_point_a(p5, p3, p2, k5)

        cv2.line(frame1, (int(pd[0]), int(pd[1])), (int(pnd[0]), int(pnd[1])), (255, 0, 0), thickness=5)
        cv2.line(frame1, (int(pu[0]), int(pu[1])), (int(pnu[0]), int(pnu[1])), (0, 0, 255), thickness=5)

    else:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=None)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=None)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=None)
        cv2.imwrite('C:/D/Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
        print('No.' + str(frames))
        RowNumber += 1
        return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor

    pointsdist = ((p2[1] - p3[1]) ** 2 + (p2[0] - p3[0]) ** 2) ** 0.5
    arclength = asin(circle_y / true_rad) * 2 * true_rad
    pointsdist = pointsdist / arclength * 100
    pre_width = pointsdist

    p_tip1, p_tip2, length = GC.get_tool_length(p2, p3, centers, k5)
    length = length / (true_rad * 2) * 100
    pre_length = length
    if pointsdist != 0:
        LW_Ratio = length / pointsdist
        if LW_Ratio > 4 or LW_Ratio < 0.5:
            cv2.putText(frame1, "Strange data", (20, 140), cv2.FONT_ITALIC, 0.5,
                        (0, 255, 0))
            sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
            sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=None)
            sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=None)
            sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=None)
            RowNumber += 1
            cv2.imwrite('C:/D/Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
            print('No.' + str(frames))
            return RowNumber, None, None, None, main_tool, main_tool_coor, assist_tool_coor
            LW_Ratio = 0
    else:
        LW_Ratio = 0
    pre_LW_Ratio = LW_Ratio

    sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
    sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pointsdist)
    sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=length)
    sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=LW_Ratio)
    RowNumber += 1

    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p2[0]), int(p2[1])), (0, 0, 255), thickness=1)
    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p3[0]), int(p3[1])), (0, 0, 255), thickness=1)
    cv2.line(frame1, (int(p2[0]), int(p2[1])), (int(p3[0]), int(p3[1])), (255, 0, 0), thickness=2)
    cv2.line(frame1, (int(p5[0]), int(p5[1])), (int(p5[0]) + 900, int(p5[1] + 900 * k5)), (255, 0, 0), thickness=2)
    cv2.line(frame1, (int(p_tip1[0]), int(p_tip1[1])), (int(p_tip2[0]), int(p_tip2[1])), (0, 255, 0), thickness=2)

    cv2.imwrite('C:/D/Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
    print('No.' + str(frames))
    cv2.waitKey(0)

    return RowNumber, pre_width, pre_length, pre_LW_Ratio, main_tool, main_tool_coor, assist_tool_coor


global video_num
global start_frame
global end_frame
video_num = 41
start_frame = 860
end_frame = 1170
if __name__ == "__main__":
    # Excel setup
    tablepath = 'Clip16_AngleTest5120-5640SL.xlsx'
    Row_Number = 2
    workbook = we.write_excel_title()
    sheet = workbook.active
    prewidth = 0
    prelength = 0
    prelwratio = 0
    find_main_tool = 0
    main_tool_coordinates = (0, 0)
    assist_tool_coordinates = (0, 0)
    # ---------------------------------

    # Laparoscopic view geo parameters
    video_constant = constant.VideoConstants()
    constant_values = video_constant.num_to_constants(video_num)()

    true_rad = constant_values[0]
    inner_rad = constant_values[1]
    outer_rad = constant_values[2]
    circle_x = constant_values[3]
    circle_y = constant_values[4]
    boundaryl = constant_values[5]
    boundaryr = constant_values[6]
    boxheight = 10
    # --------------------------------

    for frames in range(start_frame, end_frame, 1):
        if video_num % 10 == 0:
            folder_name = str(10 * (video_num // 10) - 9) + '-' + str(10 * (video_num // 10))
        else:
            folder_name = str(10 * (video_num // 10) + 1) + '-' + str(10 * (video_num // 10) + 10)
        frame = cv2.imread(
            'E:/Clip' + folder_name + '/Clip' + str(video_num) + '_1M/clip' + str(video_num) + '_' + str(
                frames) + 'M.jpg')
        frame1 = cv2.imread(
            'E:/Clip' + folder_name + '/Clip' + str(video_num) + '_1D/clip' + str(video_num) + '_' + str(
                frames) + 'D.jpg')
        mask_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        ret, contours_frame = cv2.threshold(mask_frame, 127, 255, cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(contours_frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        Row_Number, prewidth, prelength, prelwratio, find_main_tool, main_tool_coordinates, assist_tool_coordinates = calculate_distances_test(
            contours, Row_Number, prewidth, prelength,
            prelwratio, find_main_tool, main_tool_coordinates, assist_tool_coordinates)

    workbook.save(tablepath)
    print('Finish')
