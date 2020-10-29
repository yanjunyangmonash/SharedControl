import cv2
import math
import openpyxl
import sympy
from math import atan2, tan
import numpy as np

def getOrientation(pts):
    sz = len(pts)
    data_pts = np.empty((sz, 2), dtype=np.float64)
    for i in range(data_pts.shape[0]):
        data_pts[i, 0] = pts[i, 0, 0]
        data_pts[i, 1] = pts[i, 0, 1]
    # Perform PCA analysis
    mean = np.empty((0))
    mean, eigenvectors, eigenvalues = cv2.PCACompute2(data_pts, mean)
    angle = atan2(eigenvectors[0, 1], eigenvectors[0, 0])  # orientation in radians

    return angle

tablepath = 'Clip24_AngleTest99-1485.xlsx'
RowNumber = 8
ColumnNumber = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=ColumnNumber, value='Img')
sheet.cell(row=1, column=ColumnNumber+1, value='AngleValue')
sheet.cell(row=1, column=ColumnNumber+2, value='angle*distance')
sheet.cell(row=1, column=ColumnNumber+3, value='angle/distance')
sheet.cell(row=1, column=ColumnNumber+4, value='distance')
sheet.cell(row=1, column=ColumnNumber+5, value='width')
sheet.cell(row=1, column=ColumnNumber+6, value='distance ratio')
sheet.cell(row=1, column=ColumnNumber+7, value='multiply distances')

'''
frame = cv2.imread('Clip16_1M/clip16' + '_' + str(6286) + 'M.jpg')
frame1 = cv2.imread('Clip16_1D/clip16' + '_' + str(6286) + 'D.jpg')
frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
ret, frame = cv2.threshold(frame, 127, 255, cv2.THRESH_BINARY)
contours, hierarchy = cv2.findContours(frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

number = []
x_matrix = []
y_matrix = []
curve_pathx = []
curve_pathy = []
upbox_x = []
upbox_y = []
downbox_x = []
downbox_y = []

for num_of_contours in range(len(contours)):
    # Define Area
    new_contours = []
    points_touch_edge = 0
    points_touch_another_edge = 0
    # Finish parameters define

    M = cv2.moments(contours[num_of_contours], 0)
    if M['m00']:
        mass_centres_x = int(M['m10'] / M['m00'])
        mass_centres_y = int(M['m01'] / M['m00'])
    else:
        continue
    area = cv2.contourArea(contours[num_of_contours])

    if mass_centres_x >= 507:
        number.append(num_of_contours)

max_num = max(number)
number = len(contours[max_num])

for i in range(number):
    x_matrix.append(contours[max_num][i][0][0])
    y_matrix.append(contours[max_num][i][0][1])

    dist = (contours[max_num][i][0][0] - 506) ** 2 + (
            contours[max_num][i][0][1] - 279) ** 2
    if 410 ** 2 <= dist <= 430 ** 2:
        curve_pathx.append(contours[max_num][i][0][0])
        curve_pathy.append(contours[max_num][i][0][1])
    if (196 < contours[max_num][i][0][0] < 818) and (0 <= contours[max_num][i][0][1] <= 8):
        upbox_x.append(contours[max_num][i][0][0])
        upbox_y.append(contours[max_num][i][0][1])
    if (196 < contours[max_num][i][0][0] < 818) and (532 <= contours[max_num][i][0][1] <= 540):
        downbox_x.append(contours[max_num][i][0][0])
        downbox_y.append(contours[max_num][i][0][1])

p1_index = x_matrix.index(min(x_matrix))
p1 = (x_matrix[p1_index], y_matrix[p1_index])
if len(downbox_x):
    p2_index = downbox_x.index(min(downbox_x))
    p2 = (downbox_x[p2_index], downbox_y[p2_index])
    p3_index = curve_pathy.index(min(curve_pathy))
    p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
elif len(upbox_x):
    p2_index = upbox_x.index(min(upbox_x))
    p2 = (upbox_x[p2_index], upbox_y[p2_index])
    p3_index = curve_pathy.index(max(curve_pathy))
    p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
else:
    p2_index = curve_pathy.index(min(curve_pathy))
    p2 = (curve_pathx[p2_index], curve_pathy[p2_index])
    p3_index = curve_pathy.index(max(curve_pathy))
    p3 = (curve_pathx[p3_index], curve_pathy[p3_index])

cv2.line(frame1, p1, p2, (0, 0, 255), thickness=1)
cv2.line(frame1, p1, p3, (0, 0, 255), thickness=1)
ang = -math.degrees(math.atan2(p2[1]-p1[1], p2[0]-p1[0]) - math.atan2(p3[1]-p1[1], p3[0]-p1[0]))
cv2.putText(frame1, "Angle: {:.2f}".format(ang), (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
print(-ang)

cv2.imshow('test', frame1)
cv2.waitKey(0)
'''
pre_ang = 0
pre_dist = 0
pre_distl = 0
pre_distratio = 0
pre_angxd = 0
pre_ang_divide_d = 0
pre_muldist = 0

#Clip24
circle_x = 494
circle_y = 269
inner_rad = 320
outer_rad = 340
boundaryl = 304
boundaryr = 684
boxheight = 10


for frames in range(114, 1485, 1):
    frame = cv2.imread('Clip24_1M/clip24' + '_' + str(frames) + 'M.jpg')
    frame1 = cv2.imread('Clip24_1D/clip24' + '_' + str(frames) + 'D.jpg')
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(frame, 50, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    contours, _ = cv2.findContours(bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    #ret, frame = cv2.threshold(frame, 127, 255, cv2.THRESH_BINARY)
    #contours, hierarchy = cv2.findContours(frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    number = []
    x_matrix = []
    y_matrix = []
    curve_pathx = []
    curve_pathy = []
    upbox_x = []
    upbox_y = []
    downbox_x = []
    downbox_y = []
    areas = []
    p3 = (0, 0)
    twotools = 0
    true_mass_xs = []
    true_mass_ys = []
    PointsArray = []

    for num_of_contours in range(len(contours)):
        M = cv2.moments(contours[num_of_contours], 0)
        if M['m00']:
            mass_centres_x = int(M['m10'] / M['m00'])
            mass_centres_y = int(M['m01'] / M['m00'])
        else:
            continue
        area = cv2.contourArea(contours[num_of_contours])

        if mass_centres_x < circle_x:
            numbers = len(contours[num_of_contours])
            for i in range(numbers):
                if contours[num_of_contours][i][0][0] > boundaryr:
                    dist = (contours[num_of_contours][i][0][0] - circle_x) ** 2 + (
                            contours[num_of_contours][i][0][1] - circle_y) ** 2
                    if inner_rad ** 2 <= dist <= outer_rad ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('clip24_' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_ang)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_angxd)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_ang_divide_d)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 4, value=pre_distl/10)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 5, value=pre_dist/10)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 6, value=pre_distratio)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 7, value=pre_muldist)
                        RowNumber += 1
                        twotools = 1
                        break

        elif mass_centres_x >= circle_x:
            number.append(num_of_contours)
            areas.append(area)
            true_mass_xs.append(mass_centres_x)
            true_mass_ys.append(mass_centres_y)

            numbers = len(contours[num_of_contours])
            for i in range(numbers):
                PointsArray.append([[contours[num_of_contours][i][0][0], contours[num_of_contours][i][0][1]]])

                if contours[num_of_contours][i][0][0] < boundaryl:
                    dist = (contours[num_of_contours][i][0][0] - circle_x) ** 2 + (
                            contours[num_of_contours][i][0][1] - circle_y) ** 2
                    if inner_rad ** 2 <= dist <= outer_rad ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('clip24_' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_ang)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_angxd)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_ang_divide_d)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 4, value=pre_distl/10)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 5, value=pre_dist/10)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 6, value=pre_distratio)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 7, value=pre_muldist)
                        RowNumber += 1
                        twotools = 1
                        break

    if twotools == 1:
        twotools = 0
        continue

    if len(areas) == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('clip24_' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 4, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 5, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 6, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 7, value=0)
        RowNumber += 1
        continue


    max_num = areas.index(max(areas))
    true_mass_x = true_mass_xs[max_num]
    true_mass_y = true_mass_ys[max_num]
    max_num = number[max_num]
    number = len(contours[max_num])

    for i in range(number):
        x_matrix.append(contours[max_num][i][0][0])
        y_matrix.append(contours[max_num][i][0][1])

        dist = (contours[max_num][i][0][0] - circle_x) ** 2 + (
                contours[max_num][i][0][1] - circle_y) ** 2
        if inner_rad ** 2 <= dist <= outer_rad ** 2:
            curve_pathx.append(contours[max_num][i][0][0])
            curve_pathy.append(contours[max_num][i][0][1])
        if (boundaryl+15 < contours[max_num][i][0][0] < boundaryr-15) and (0 <= contours[max_num][i][0][1] <= boxheight):
            upbox_x.append(contours[max_num][i][0][0])
            upbox_y.append(contours[max_num][i][0][1])
        if (boundaryl+15 < contours[max_num][i][0][0] < boundaryr-15) and (540 - boxheight <= contours[max_num][i][0][1] <= 540):
            downbox_x.append(contours[max_num][i][0][0])
            downbox_y.append(contours[max_num][i][0][1])

    p1_index = x_matrix.index(min(x_matrix))
    #p1_index = 0
    #pre_min = outer_rad
    #for i in range(len(x_matrix)):
        #distance = ((x_matrix[i] - circle_x) ** 2 + (y_matrix[i] - circle_y) ** 2) ** 0.5
        #if distance < pre_min:
            #pre_min = distance
            #p1_index = i
    p1 = (x_matrix[p1_index], y_matrix[p1_index])
    p4 = (true_mass_x, true_mass_y)
    PointsArray = np.array(PointsArray)
    tool_angle = getOrientation(PointsArray)
    tool_k = tan(tool_angle)

    if len(downbox_x) and len(curve_pathx):
        #p2_index = downbox_x.index(min(downbox_x))
        #p2 = (downbox_x[p2_index], downbox_y[p2_index])
        p3_index = curve_pathy.index(min(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2x = ((1-tool_k**2)*p3[0]-2*tool_k*(-1)*p3[1]-2*tool_k*(p4[1]-tool_k*p4[0]))/(1+tool_k**2)
        p2y = ((tool_k**2-1)*p3[1]-2*tool_k*(-1)*p3[0]-2*(-1)*(p4[1]-tool_k*p4[0]))/(1+tool_k**2)
        p2 = (int(p2x), int(p2y))
    elif len(downbox_x) and len(curve_pathx) == 0:
        p2_index = downbox_x.index(min(downbox_x))
        p2 = (downbox_x[p2_index], downbox_y[p2_index])
        p3_index = downbox_x.index(max(downbox_x))
        p3 = (downbox_x[p3_index], downbox_y[p3_index])
        p4p2dist = ((p2[1] - p4[1]) ** 2 + (p2[0] - p4[0]) ** 2) ** 0.5
        p4p3dist = ((p3[1] - p4[1]) ** 2 + (p3[0] - p4[0]) ** 2) ** 0.5
        if p4p3dist > p4p2dist:
            p2x = ((1 - tool_k ** 2) * p3[0] - 2 * tool_k * (-1) * p3[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2y = ((tool_k ** 2 - 1) * p3[1] - 2 * tool_k * (-1) * p3[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2 = (int(p2x), int(p2y))
        else:
            p3x = ((1-tool_k**2)*p2[0]-2*tool_k*(-1)*p2[1]-2*tool_k*(p4[1]-tool_k*p4[0]))/(1+tool_k**2)
            p3y = ((tool_k**2-1)*p2[1]-2*tool_k*(-1)*p2[0]-2*(-1)*(p4[1]-tool_k*p4[0]))/(1+tool_k**2)
            p3 = (int(p3x), int(p3y))
    elif len(upbox_x) and len(curve_pathx):
        #p2_index = upbox_x.index(min(upbox_x))
        #p2 = (upbox_x[p2_index], upbox_y[p2_index])
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p2x = ((1 - tool_k ** 2) * p3[0] - 2 * tool_k * (-1) * p3[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                    1 + tool_k ** 2)
        p2y = ((tool_k ** 2 - 1) * p3[1] - 2 * tool_k * (-1) * p3[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                    1 + tool_k ** 2)
        p2 = (int(p2x), int(p2y))
    elif len(upbox_x) and len(curve_pathx) == 0:
        p2_index = upbox_x.index(min(upbox_x))
        p2 = (upbox_x[p2_index], upbox_y[p2_index])
        p3_index = upbox_x.index(max(upbox_x))
        p3 = (upbox_x[p3_index], upbox_y[p3_index])
        p4p2dist = ((p2[1] - p4[1]) ** 2 + (p2[0] - p4[0]) ** 2) ** 0.5
        p4p3dist = ((p3[1] - p4[1]) ** 2 + (p3[0] - p4[0]) ** 2) ** 0.5
        if p4p3dist > p4p2dist:
            p2x = ((1 - tool_k ** 2) * p3[0] - 2 * tool_k * (-1) * p3[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2y = ((tool_k ** 2 - 1) * p3[1] - 2 * tool_k * (-1) * p3[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2 = (int(p2x), int(p2y))
        else:
            p3x = ((1 - tool_k ** 2) * p2[0] - 2 * tool_k * (-1) * p2[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p3y = ((tool_k ** 2 - 1) * p2[1] - 2 * tool_k * (-1) * p2[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p3 = (int(p3x), int(p3y))
    elif len(curve_pathx):
        p2_index = curve_pathy.index(min(curve_pathy))
        p2 = (curve_pathx[p2_index], curve_pathy[p2_index])
        p3_index = curve_pathy.index(max(curve_pathy))
        p3 = (curve_pathx[p3_index], curve_pathy[p3_index])
        p4p2dist = ((p2[1] - p4[1]) ** 2 + (p2[0] - p4[0]) ** 2) ** 0.5
        p4p3dist = ((p3[1] - p4[1]) ** 2 + (p3[0] - p4[0]) ** 2) ** 0.5
        if p4p3dist > p4p2dist:
            p2x = ((1 - tool_k ** 2) * p3[0] - 2 * tool_k * (-1) * p3[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2y = ((tool_k ** 2 - 1) * p3[1] - 2 * tool_k * (-1) * p3[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p2 = (int(p2x), int(p2y))
        else:
            p3x = ((1 - tool_k ** 2) * p2[0] - 2 * tool_k * (-1) * p2[1] - 2 * tool_k * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p3y = ((tool_k ** 2 - 1) * p2[1] - 2 * tool_k * (-1) * p2[0] - 2 * (-1) * (p4[1] - tool_k * p4[0])) / (
                        1 + tool_k ** 2)
            p3 = (int(p3x), int(p3y))

    if p3 != (0, 0):
        pointsdist = ((p2[1] - p3[1]) ** 2 + (p2[0] - p3[0]) ** 2) ** 0.5
        if p2[0] > p3[0]:
            p5 = (int(p3[0]+(p2[0] - p3[0])/2), int(p3[1]+(p2[1] - p3[1])/2))
        else:
            p5 = (int(p2[0] + (p3[0] - p2[0]) / 2), int(p2[1] + (p3[1] - p2[1]) / 2))
        #p1p2dist = ((p2[1] - p4[1]) ** 2 + (p2[0] - p4[0]) ** 2) ** 0.5
        #p1p3dist = ((p3[1] - p4[1]) ** 2 + (p3[0] - p4[0]) ** 2) ** 0.5
        #if p1p2dist > p1p3dist:
            #x = sympy.Symbol('x')
            #y = sympy.Symbol('y')
            #solved_value = sympy.solve(
                #[(p3[1] - y) / (p3[0] - x) - (p3[1] - p4[1]) / (p3[0] - p4[0]),
                 #((y - p4[1]) ** 2 + (x - p4[0]) ** 2) - p1p2dist ** 2], [x, y])
            #for i in range(len(solved_value)):
                #if (solved_value[i][0] - p3[0]) * (p3[0] - p4[0]) > 0:
                    #p3 = (int(solved_value[i][0]), int(solved_value[i][1]))
                    #break
            '''
            solved_value = sympy.solve(
                [(p2[1] - y) / (p2[0] - x) - (p2[1] - p1[1]) / (p2[0] - p1[0]),
                 ((y - p1[1]) ** 2 + (x - p1[0]) ** 2) - p1p3dist ** 2], [x, y])
            for i in range(len(solved_value)):
                if (p2[0] - solved_value[i][0]) * (solved_value[i][0] - p1[0]) > 0:
                    p2 = (int(solved_value[i][0]), int(solved_value[i][1]))
                    break
            '''
        #else:
            #x = sympy.Symbol('x')
            #y = sympy.Symbol('y')
            #solved_value = sympy.solve(
                #[(p2[1] - y) / (p2[0] - x) - (p2[1] - p4[1]) / (p2[0] - p4[0]),
                 #((y - p4[1]) ** 2 + (x - p4[0]) ** 2) - p1p3dist ** 2], [x, y])
            #for i in range(len(solved_value)):
                #if (solved_value[i][0] - p2[0]) * (p2[0] - p4[0]) > 0:
                    #p2 = (int(solved_value[i][0]), int(solved_value[i][1]))
                    #break
            '''
            solved_value = sympy.solve(
                [(p3[1] - y) / (p3[0] - x) - (p3[1] - p1[1]) / (p3[0] - p1[0]),
                 ((y - p1[1]) ** 2 + (x - p1[0]) ** 2) - p1p2dist ** 2], [x, y])
            for i in range(len(solved_value)):
                if (p3[0] - solved_value[i][0]) * (solved_value[i][0] - p1[0]) > 0:
                    p3 = (int(solved_value[i][0]), int(solved_value[i][1]))
                    break
            '''

        #pointsdist = ((p2[1] - p3[1]) ** 2 + (p2[0] - p3[0]) ** 2) ** 0.5
        #p1p2dist = ((p2[1] - p1[1]) ** 2 + (p2[0] - p1[0]) ** 2) ** 0.5
        #p1p3dist = ((p3[1] - p1[1]) ** 2 + (p3[0] - p1[0]) ** 2) ** 0.5
        #print(p1p2dist)
        #print(p1p3dist)
        #if p2[0] > p3[0]:
            #p5 = (int(p3[0]+(p2[0] - p3[0])/2), int(p3[1]+(p2[1] - p3[1])/2))
        #else:
            #p5 = (int(p2[0] + (p3[0] - p2[0]) / 2), int(p2[1] + (p3[1] - p2[1]) / 2))


        cv2.line(frame1, p4, p2, (0, 0, 255), thickness=1)
        cv2.line(frame1, p5, p2, (0, 0, 255), thickness=1)
        cv2.line(frame1, p4, p3, (0, 0, 255), thickness=1)
        cv2.line(frame1, p5, p3, (0, 0, 255), thickness=1)
        cv2.line(frame1, p4, p5, (0, 0, 255), thickness=1)

        ang = abs(math.degrees(math.atan2(p2[1] - p1[1], p2[0] - p1[0]) - math.atan2(p3[1] - p1[1], p3[0] - p1[0])))
        angxd = ang * pointsdist / 100
        ang_divide_d = ang / pointsdist*100
        longdist = ((p1[1] - p4[1]) ** 2 + (p1[0] - p4[0]) ** 2) ** 0.5
        distrato = longdist / pointsdist * 10
        muldist = longdist * pointsdist / 2000
    else:
        ang = 0

    pre_ang = ang
    pre_dist = pointsdist
    pre_ang_divide_d = ang_divide_d
    pre_angxd = angxd
    pre_distl = longdist
    pre_distratio = distrato
    pre_muldist = muldist
    cv2.putText(frame1, "Angle: {:.2f}".format(ang), (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "angxd: {:.2f}".format(angxd), (20, 40), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "ang/d: {:.2f}".format(ang_divide_d), (20, 60), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "distance: {:.2f}".format(longdist), (20, 80), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "width: {:.2f}".format(pointsdist), (20, 100), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "distrato: {:.2f}".format(distrato), (20, 120), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
    cv2.putText(frame1, "muldist: {:.2f}".format(muldist), (20, 140), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

    cv2.imwrite('Clip24_GUITest/clip24' + '_' + str(frames) + '.jpg', frame1)
    cv2.waitKey(0)
    sheet.cell(row=RowNumber, column=ColumnNumber, value=('clip24_' + str(frames)))
    sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=ang)
    sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=angxd)
    sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=ang_divide_d)
    sheet.cell(row=RowNumber, column=ColumnNumber + 4, value=longdist/10)
    sheet.cell(row=RowNumber, column=ColumnNumber + 5, value=pointsdist/10)
    sheet.cell(row=RowNumber, column=ColumnNumber + 6, value=distrato)
    sheet.cell(row=RowNumber, column=ColumnNumber + 7, value=muldist)
    RowNumber += 1
    print('clip24_' + str(frames))

workbook.save(tablepath)
print('Finish')


