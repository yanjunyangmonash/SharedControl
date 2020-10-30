import cv2
import openpyxl
from math import atan, tan, asin
import numpy as np
import constant

def circle_line_segment_intersection(circle_center, circle_radius, pt1, pt2, full_line=True, tangent_tol=1e-9):
    """ Find the points at which a circle intersects a line-segment.  This can happen at 0, 1, or 2 points.

    :param circle_center: The (x, y) location of the circle center
    :param circle_radius: The radius of the circle
    :param pt1: The (x, y) location of the first point of the segment
    :param pt2: The (x, y) location of the second point of the segment
    :param full_line: True to find intersections along full line - not just in the segment.  False will just return intersections within the segment.
    :param tangent_tol: Numerical tolerance at which we decide the intersections are close enough to consider it a tangent
    :return Sequence[Tuple[float, float]]: A list of length 0, 1, or 2, where each element is a point at which the circle intercepts a line segment.

    Note: We follow: http://mathworld.wolfram.com/Circle-LineIntersection.html
    """

    (p1x, p1y), (p2x, p2y), (cx, cy) = pt1, pt2, circle_center
    (x1, y1), (x2, y2) = (p1x - cx, p1y - cy), (p2x - cx, p2y - cy)
    dx, dy = (x2 - x1), (y2 - y1)
    dr = (dx ** 2 + dy ** 2)**.5
    big_d = x1 * y2 - x2 * y1
    discriminant = circle_radius ** 2 * dr ** 2 - big_d ** 2

    if discriminant < 0:  # No intersection between circle and line
        return []
    else:  # There may be 0, 1, or 2 intersections with the segment
        intersections = [
            (cx + (big_d * dy + sign * (-1 if dy < 0 else 1) * dx * discriminant**.5) / dr ** 2,
             cy + (-big_d * dx + sign * abs(dy) * discriminant**.5) / dr ** 2)
            for sign in ((1, -1) if dy < 0 else (-1, 1))]  # This makes sure the order along the segment is correct
        if not full_line:  # If only considering the segment, filter out intersections that do not fall within the segment
            fraction_along_segment = [(xi - p1x) / dx if abs(dx) > abs(dy) else (yi - p1y) / dy for xi, yi in intersections]
            intersections = [pt for pt, frac in zip(intersections, fraction_along_segment) if 0 <= frac <= 1]
        if len(intersections) == 2 and abs(discriminant) <= tangent_tol:  # If line is tangent to circle, return just one point (as both intersections have same location)
            return [intersections[0]]
        else:
            return intersections

def calculate_distances(no_of_frame, contours_number, circle_x_coor, circle_y_coor, left_bound, right_bound, excel_number, inner_r, outer_r, real_r, box_height, pre_width, pre_length, pre_LW_Ratio):
    # Set up parameters
    contour_areas = []
    two_tools_touch = 0
    true_mass_xs = []
    true_mass_ys = []
    number = []
    RowNumber = excel_number
    ColumnNumber = 1
    seperate_end_effector_mask = 0

    # Manually set metrics (Use Clip33 as the ref)
    # For two masks classification
    area_ratio_metrics = 10
    mask_dist_metrics = (real_r * 2) * 0.374
    end_effector_detail = (real_r * 2) * 0.18
    tool_body_concave = (real_r * 2) * 0.0449
    k_ratio = 0.3
    length_ratio_metrics = 50
    small_area_metrics = (np.pi * true_rad * true_rad) * 0.005707


    for num_of_contours in range(len(contours_number)):
        M = cv2.moments(contours_number[num_of_contours], 0)
        if M['m00']:
            mass_centres_x = int(M['m10'] / M['m00'])
            mass_centres_y = int(M['m01'] / M['m00'])
        else:
            number.append(0)
            contour_areas.append(0)
            true_mass_xs.append(0)
            true_mass_ys.append(0)
            continue

        if mass_centres_x < left_bound:
            number.append(0)
            contour_areas.append(0)
            true_mass_xs.append(0)
            true_mass_ys.append(0)
            numbers = len(contours_number[num_of_contours])
            for i in range(numbers):
                if contours_number[num_of_contours][i][0][0] > right_bound:
                    dist = (contours_number[num_of_contours][i][0][0] - circle_x) ** 2 + (
                            contours_number[num_of_contours][i][0][1] - circle_y) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_rad ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(no_of_frame)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        two_tools_touch = 1
                        break

        elif left_bound <= mass_centres_x < circle_x_coor:
            right_tool = 0
            numbers = len(contours_number[num_of_contours])
            for j in range(numbers):
                if contours_number[num_of_contours][j][0][1] > 525 or contours_number[num_of_contours][j][0][1] < 15:
                    right_tool = 1

                if contours_number[num_of_contours][j][0][0] > right_bound:
                    dist = (contours_number[num_of_contours][j][0][0] - circle_x_coor) ** 2 + (
                            contours_number[num_of_contours][j][0][1] - circle_y_coor) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_r ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        two_tools_touch = 1
                        break
                elif contours_number[num_of_contours][j][0][0] < left_bound:
                    dist = (contours_number[num_of_contours][j][0][0] - circle_x_coor) ** 2 + (
                            contours_number[num_of_contours][j][0][1] - circle_y_coor) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_r ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        right_tool = 0
                        break

            # Calculate min area rect
            if right_tool:
                rect = cv2.minAreaRect(contours_number[num_of_contours])
                box = cv2.boxPoints(rect)
                box_h = abs(((box[0][0] - box[1][0]) ** 2 + (box[0][1] - box[1][1])) ** 0.5)
                box_w = abs(((box[1][0] - box[2][0]) ** 2 + (box[1][1] - box[2][1])) ** 0.5)
                number.append(num_of_contours)
                true_mass_xs.append(mass_centres_x)
                true_mass_ys.append(mass_centres_y)

                if box_h*box_w > small_area_metrics:
                    contour_areas.append(box_h*box_w)
                else:
                    contour_areas.append(0)
            else:
                number.append(0)
                contour_areas.append(0)
                true_mass_xs.append(0)
                true_mass_ys.append(0)


        elif mass_centres_x >= circle_x_coor:
            # Calculate min area rect
            rect = cv2.minAreaRect(contours_number[num_of_contours])
            box = cv2.boxPoints(rect)
            box_h = abs(((box[0][0] - box[1][0]) ** 2 + (box[0][1] - box[1][1])) ** 0.5)
            box_w = abs(((box[1][0] - box[2][0]) ** 2 + (box[1][1] - box[2][1])) ** 0.5)
            number.append(num_of_contours)
            contour_areas.append(box_h*box_w)
            true_mass_xs.append(mass_centres_x)
            true_mass_ys.append(mass_centres_y)

            numbers = len(contours_number[num_of_contours])
            for l in range(numbers):
                if contours_number[num_of_contours][l][0][0] < left_bound:
                    dist = (contours_number[num_of_contours][l][0][0] - circle_x_coor) ** 2 + (
                            contours_number[num_of_contours][l][0][1] - circle_y_coor) ** 2
                    if (inner_rad) ** 2 <= dist <= outer_r ** 2:
                        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
                        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=pre_width)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=pre_length)
                        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=pre_LW_Ratio)
                        RowNumber += 1
                        two_tools_touch = 1
                        break


    if two_tools_touch == 1:
        two_tools_touch = 0
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        cv2.putText(frame1, "Two tools contact together", (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
        cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
        print('No.' + str(frames))
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    if len(contour_areas) == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        cv2.putText(frame1, "No Tools on the right", (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
        cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
        print('No.' + str(frames))
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    if max(contour_areas) < 300 or len(contour_areas) == 0:
        sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
        sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
        RowNumber += 1
        cv2.putText(frame1, "Bad mask situation A", (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
        cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
        print('No.' + str(frames))
        return RowNumber, pre_width, pre_length, pre_LW_Ratio

    # Find max contour on the right side
    sorted_contour_areas = sorted(contour_areas)
    max_num_id = contour_areas.index(max(contour_areas))
    cv2.drawContours(frame1, contours_number[max_num_id], -1, (255, 0, 0), 3)
    if len(contour_areas) > 1 and sorted_contour_areas[-2] != 0:
        sec_max_num_id = contour_areas.index(sorted_contour_areas[-2])
        cv2.drawContours(frame1, contours_number[sec_max_num_id], -1, (0, 0, 255), 3)
        area_ratio = sorted_contour_areas[-2]/max(contour_areas)*100
        cv2.putText(frame1, "Area Ratio: {:.2f}%".format(area_ratio), (20, 20), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

        max_x = true_mass_xs[max_num_id]
        max_y = true_mass_ys[max_num_id]
        sec_max_x = true_mass_xs[sec_max_num_id]
        sec_max_y = true_mass_ys[sec_max_num_id]
        mask_dist = ((max_x - sec_max_x) ** 2 + (max_y - sec_max_y) ** 2) ** 0.5

        cv2.circle(frame1, (int(max_x), int(max_y)), 8, (0, 0, 255), 5)
        cv2.circle(frame1, (int(sec_max_x), int(sec_max_y)), 8, (255, 0, 0), 5)


        if area_ratio > area_ratio_metrics and mask_dist < mask_dist_metrics:
            cv2.putText(frame1, "Only have end effector", (20, 40), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
            seperate_end_effector_mask = 1

    true_mass_x = true_mass_xs[max_num_id]
    true_mass_y = true_mass_ys[max_num_id]
    max_num = number[max_num_id]

    if seperate_end_effector_mask == 0:
        hull = cv2.convexHull(contours_number[max_num_id], clockwise=False, returnPoints=False)
        try:
            defects = cv2.convexityDefects(contours_number[max_num_id], hull)
        except cv2.error as e:
            sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
            sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=None)
            sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=None)
            sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=None)
            RowNumber += 1
            cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
            return RowNumber, None, None, None

        far_list = []

        for j in range(defects.shape[0]):
            s, e, f, d = defects[j, 0]
            point_dist = np.int(d)
            far_list.append(point_dist)

        new_list = sorted(far_list)
        max_number = far_list.index(new_list[-1])
        s, e, f, d = defects[max_number, 0]
        far = tuple(contours_number[max_num_id][f][0])
        start = tuple(contours_number[max_num_id][s][0])
        end = tuple(contours_number[max_num_id][e][0])
        point_dist = np.int(d)
        cv2.line(frame1, start, end, (0, 225, 0), 2)
        cv2.circle(frame1, far, 8, (0, 255, 0), 5)

        cv2.putText(frame1, "Farthest dist: {:.2f}".format(point_dist/256), (20, 60), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

        if (point_dist/256) < end_effector_detail:
            if (point_dist/256) < tool_body_concave:
                cv2.putText(frame1, "Record", (20, 80), cv2.FONT_ITALIC, 0.5, (0, 255, 0))
            else:
                cv2.circle(frame1, start, 8, (255, 255, 0), 5)
                cv2.circle(frame1, end, 8, (0, 255, 0), 5)
                cv2.putText(frame1, "Consider", (20, 80), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

                k_start = abs((far[1] - start[1]) / (far[0] - start[0]))
                k_end = abs((far[1]-end[1])/(far[0]-end[0]))

                if k_start > k_end:
                    if k_end/k_start < k_ratio:
                        pt1 = end
                        pt2 = far
                    else:
                        pt1 = ((end[0]+start[0])/2, (end[1]+start[1])/2)
                        pt2 = far
                else:
                    if k_start/k_end < k_ratio:
                        pt1 = start
                        pt2 = far
                    else:
                        pt1 = ((end[0]+start[0])/2, (end[1]+start[1])/2)
                        pt2 = far
                far_to_effector = ((far[0] - pt1[0]) ** 2 + (far[1] - pt1[1]) ** 2) ** 0.5

                circle_cen = (circle_x, circle_y)
                circle_rad = true_rad
                new_point = circle_line_segment_intersection(circle_cen, circle_rad, pt1, pt2, full_line=True, tangent_tol=1e-9)
                if new_point[0][0] > new_point[1][0]:
                    new_point = (new_point[0][0], new_point[0][1])
                else:
                    new_point = (new_point[1][0], new_point[1][1])
                cv2.line(frame1, far, (int(new_point[0]), int(new_point[1])), (0, 0, 225), 2)

                far_to_edge = ((far[0]-new_point[0])**2+(far[1]-new_point[1])**2)**0.5
                length_ratio = far_to_effector/far_to_edge*100
                cv2.putText(frame1, "effector length ratio: {:.2f}".format(length_ratio), (20, 100), cv2.FONT_ITALIC, 0.5,
                            (0, 255, 0))
                if length_ratio <= length_ratio_metrics:
                    cv2.putText(frame1, "Record", (20, 120), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

        else:
            cv2.putText(frame1, "Dont record", (20, 80), cv2.FONT_ITALIC, 0.5, (0, 255, 0))

    sheet.cell(row=RowNumber, column=ColumnNumber, value=('No.' + str(frames)))
    sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=0)
    sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=0)
    sheet.cell(row=RowNumber, column=ColumnNumber + 3, value=0)
    RowNumber += 1
    cv2.imwrite('Clip16SL/clip16' + '_' + str(frames) + '.jpg', frame1)
    print('No.' + str(frames))
    cv2.waitKey(0)

    return RowNumber, pre_width, pre_length, pre_LW_Ratio

if __name__ == "__main__":
    # Excel setup
    tablepath = 'Clip16_AngleTest5120-5640SL.xlsx'
    Row_Number = 2
    ColumnNumber = 1
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=ColumnNumber, value='Img')
    sheet.cell(row=1, column=ColumnNumber + 1, value='Width')
    sheet.cell(row=1, column=ColumnNumber + 2, value='Length')
    sheet.cell(row=1, column=ColumnNumber + 3, value='LW_Ratio')
    prewidth = 0
    prelength = 0
    prelwratio = 0
    # ---------------------------------

    # Laparoscopic view geo parameters
    video_num = 42
    video_constant = constant.VideoConstants()
    constant_values = video_constant.num_to_constants(video_num)()

    true_rad = constant_values[0]
    inner_rad = constant_values[1]
    outer_rad = constant_values[2]
    circle_x = constant_values[3]
    circle_y = constant_values[4]
    boundaryl = constant_values[5]
    boundaryr = constant_values[6]
    boxheight = 10
    # --------------------------------

    for frames in range(6120, 6125, 1):
        if video_num % 10 == 0:
            folder_name = str(10 * (video_num // 10) - 9) + '-' + str(10 * (video_num // 10))
        else:
            folder_name = str(10*(video_num//10)+1) + '-' + str(10*(video_num//10)+10)
        frame = cv2.imread('D:/Clip' + folder_name + '/Clip' + str(video_num) + '_1M/clip' + str(video_num) + '_' + str(
            frames) + 'M.jpg')
        frame1 = cv2.imread(
            'D:/Clip' + folder_name + '/Clip' + str(video_num) + '_1D/clip' + str(video_num) + '_' + str(
                frames) + 'D.jpg')
        mask_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        ret, contours_frame = cv2.threshold(mask_frame, 127, 255, cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(contours_frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        Row_Number, prewidth, prelength, prelwratio = calculate_distances(frames, contours, circle_x, circle_y,
                                                                          boundaryl, boundaryr,
                                                                          Row_Number, inner_rad, outer_rad, true_rad,
                                                                          boxheight, prewidth, prelength, prelwratio)

    workbook.save(tablepath)
    print('Finish')
