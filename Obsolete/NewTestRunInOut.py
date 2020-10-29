import openpyxl
import cv2
import math


def draw_defined_boundaries(frame, circle_center_coor, two_radius, box_size, color=(0, 0, 255)):
    # Draw the boundary for edge-contacted tool detection
    inner_radius = two_radius[0]
    outer_radius = two_radius[1]

    cv2.circle(frame, circle_center_coor, inner_radius, color, thickness=1)
    cv2.circle(frame, circle_center_coor, outer_radius, color, thickness=1)

    # Extra height is used to narrow the distance between two boundaries
    defined_height = box_size[2]
    left_boundary_x = box_size[0]
    right_boundary_x = box_size[1]
    upper_box_points = [(left_boundary_x, 0), (right_boundary_x, defined_height)]
    bottom_box_points = [(left_boundary_x, 540 - defined_height), (right_boundary_x, 540)]

    cv2.rectangle(frame, upper_box_points[0], upper_box_points[1], color, thickness=1)
    cv2.rectangle(frame, bottom_box_points[0], bottom_box_points[1], color, thickness=1)

    center_line_x = circle_center_coor[0]
    center_line_points = [(center_line_x, 0), (center_line_x, 540)]
    cv2.line(frame, center_line_points[0], center_line_points[1], color, thickness=1)


def detect_tools_contact(contour_group, contour_number, num_of_contour_points, circle_center_coor, in_radius,
                         out_radius):
    distance = (contour_group[contour_number][num_of_contour_points][0][0] - circle_center_coor[0]) ** 2 + (
            contour_group[contour_number][num_of_contour_points][0][1] - circle_center_coor[1]) ** 2

    if in_radius ** 2 <= distance <= out_radius ** 2:
        return 1
    else:
        return None


def detect_tool_from_LR(contour_group, contour_number, num_of_contour_points, circle_center_coor, in_radius,
                        out_radius):
    distance = (contour_group[contour_number][num_of_contour_points][0][0] - circle_center_coor[0]) ** 2 + (
            contour_group[contour_number][num_of_contour_points][0][1] - circle_center_coor[1]) ** 2

    if in_radius ** 2 <= distance <= out_radius ** 2:
        return 1, [[contour_group[contour_number][num_of_contour_points][0][0],
                    contour_group[contour_number][num_of_contour_points][0][1]]]
    else:
        return None, None


def detect_tool_from_UD(contour_group, contour_number, num_of_contour_points, boundary_L, boundary_R, zone_height):
    if boundary_L < contour_group[contour_number][num_of_contour_points][0][0] < \
            boundary_R and 0 <= contour_group[contour_number][num_of_contour_points][0][1] <= zone_height:
        return 1, [[contour_group[contour_number][num_of_contour_points][0][0],
                    contour_group[contour_number][num_of_contour_points][0][1]]]

    elif boundary_L < contour_group[contour_number][num_of_contour_points][0][0] < \
            boundary_R and 540 - zone_height <= contour_group[contour_number][num_of_contour_points][0][1] <= 540:
        return 1, [[contour_group[contour_number][num_of_contour_points][0][0],
                    contour_group[contour_number][num_of_contour_points][0][1]]]

    else:
        return None, None


def contour_points_in_edges(frame, points_detect_tools, points_detect_single_tool, area_of_tool, mass_centres_coor,
                            selected_contours):
    twoConnectedTool_area_sets = 0
    singleTool = 0

    if points_detect_tools >= 3:
        twoConnectedTool_area_sets = area_of_tool
        cv2.circle(frame, mass_centres_coor, radius=4, color=(0, 0, 255), thickness=4)

    # Use 5 for clip 16
    elif points_detect_single_tool >= 3:
        # Assign Area
        singleTool = area_of_tool
        # Yellow circle for the mass centre (Recognise as a tool, example 2570)
        cv2.circle(frame, mass_centres_coor, radius=4, color=(0, 255, 255), thickness=4)
        # Green points shown if the tool touch with the lens boundary
        for k in range(len(selected_contours)):
            cv2.circle(frame, (selected_contours[k][0][0], selected_contours[k][0][1]), radius=1, color=(0, 255, 0),
                       thickness=4)
    return twoConnectedTool_area_sets, singleTool


def tool_or_noise_l(leftTool_groups, pre_leftToolTot):
    if len(leftTool_groups) > 2:
        leftTool_groups.sort(reverse=True)
        leftToolTot_area = sum(leftTool_groups[0:2])
        leftTool1_area = leftTool_groups[0]
        leftTool2_area = leftTool_groups[1]

        # If the tool is already in the view
        if pre_leftToolTot:
            leftToolChange = leftToolTot_area - pre_leftToolTot
            leftchangerate = leftToolChange / pre_leftToolTot * 100

            # Prevent random noise, I assume if the change rate is too high, then there should be some noise. The value is set from image 183-185
            if abs(leftchangerate) >= 1500:
                # Just for test, the value given to the current total value should consider the moving trend(Using Kalman filter)
                leftToolTot_area = pre_leftToolTot
                # The change rate should keep the previous value if lost tracking?
                leftchangerate = 0
        # If the tool is newly inserted, 500% is signal for new tool insert?
        else:
            leftchangerate = 500

    # If no tool
    else:
        # Prevent if the detect mask contour doesn't touch the lens edge (image840-873)
        if pre_leftToolTot > 12000:
            leftToolTot_area = pre_leftToolTot
            # The change rate should keep the previous value if lost tracking?
            leftchangerate = 0
        # Try to detect normal tool withdrawal, the threshold need to be adjusted later
        elif 1000 < pre_leftToolTot < 10000:
            leftchangerate = -200
            leftToolTot_area = 0
        else:
            leftchangerate = 0
            leftToolTot_area = 0
        leftTool1_area = 0
        leftTool2_area = 0

    return leftToolTot_area, leftchangerate, leftTool1_area, leftTool2_area


def tool_or_noise_r(rightTool_groups, pre_rightToolTot, num_frozen_frames, pre_twoTools_touch, circle_rad):
    # From all contours in one frame, if find tools on the right side
    if len(rightTool_groups) > 2:
        rightTool_groups.sort(reverse=True)
        rightToolTot_area = sum(rightTool_groups[0:2])
        rightTool1_area = rightTool_groups[0]
        rightTool2_area = rightTool_groups[1]

        if pre_rightToolTot:
            rightToolChange = rightToolTot_area - pre_rightToolTot
            right_changerate = rightToolChange / pre_rightToolTot * 100

            # This seems only prevents the noise occur after already getting an accurate mask, what if the noise comes
            # before receiving a good mask? (Clip25 12-20)
            if abs(right_changerate) >= 1500:
                rightToolTot_area = pre_rightToolTot
                right_changerate = 0
        # New tool inserts
        else:
            right_changerate = 500

    else:
        Circle_threshold = 0.03 * math.pi * (circle_rad) ** 2
        # If two tools touch together, the area will keep the value before two tool contact
        if pre_twoTools_touch:
            rightToolTot_area = pre_rightToolTot
            right_changerate = 0
        # Prevent tracking lost
        # Use 12000 for clip 16
        elif pre_rightToolTot > Circle_threshold:
            rightToolTot_area = pre_rightToolTot
            right_changerate = 0
        # Detect normal tool withdrawal
        # Use 10000 for clip 16
        elif pre_rightToolTot < Circle_threshold:
            right_changerate = -200
            rightToolTot_area = 0
        # Detect no tool in the view
        #else:
            #right_changerate = - pre_rightToolTot / (pre_rightToolTot + 0.000001) * 100
            #rightToolTot_area = 0
        rightTool1_area = 0
        rightTool2_area = 0

    # If the area is frozen for 5 frames, force it to be zero
    if pre_rightToolTot > 0 and pre_rightToolTot == rightToolTot_area and pre_twoTools_touch == 0:
        num_frozen_frames += 1
        if num_frozen_frames > 4:
            rightToolTot_area = 0
            num_frozen_frames = 0

    return rightToolTot_area, right_changerate, num_frozen_frames, rightTool1_area, rightTool2_area


def identify_mask_position(frame, contours_number, center_of_circle, two_radius, box_size):
    leftTool_sets = [0, 0]
    rightTool_sets = [0, 0]
    twoConnectedTool_area_sets = [0]

    inner_radius = two_radius[0]
    outer_radius = two_radius[1]

    defined_height = box_size[2]
    left_boundary_x = box_size[0]
    right_boundary_x = box_size[1]

    for num_of_contours in range(len(contours_number)):
        # Define parameters
        new_contours = []
        points_touch_edge = 0
        points_touch_another_edge = 0
        # Finish

        M = cv2.moments(contours_number[num_of_contours], 0)
        if M['m00']:
            mass_centres_x = int(M['m10'] / M['m00'])
            mass_centres_y = int(M['m01'] / M['m00'])
            area = cv2.contourArea(contours_number[num_of_contours])
            mass_centres = (mass_centres_x, mass_centres_y)
        else:
            continue

        center_line_x = center_of_circle[0]
        if mass_centres_x <= center_line_x:
            number = len(contours_number[num_of_contours])
            for i in range(number):
                # Detect whether two tools touch together
                if contours_number[num_of_contours][i][0][0] > right_boundary_x:
                    Tools_contact = detect_tools_contact(contours_number, num_of_contours, i, center_of_circle,
                                                         inner_radius, outer_radius)
                    if Tools_contact:
                        points_touch_another_edge += Tools_contact
                        cv2.circle(frame, (contours_number[num_of_contours][i][0][0],
                                           contours_number[num_of_contours][i][0][1]), radius=1,
                                   color=(130, 100, 0), thickness=4)
                # Detect if the tool comes from the left side
                else:
                    if contours_number[num_of_contours][i][0][0] < center_line_x:
                        if contours_number[num_of_contours][i][0][0] < left_boundary_x:
                            # Blue points show tool boundary between left_boundary_line_x and left_curve_path
                            cv2.circle(frame, (contours_number[num_of_contours][i][0][0],
                                               contours_number[num_of_contours][i][0][1]),
                                       radius=1, color=(200, 210, 0), thickness=4)
                            result_tool_from_LR = detect_tool_from_LR(contours_number, num_of_contours, i,
                                                                      center_of_circle, inner_radius, outer_radius)
                            if result_tool_from_LR[0]:
                                points_touch_edge += result_tool_from_LR[0]
                                new_contours.append(result_tool_from_LR[1])

                        # Try to find tools insert from up edge or bottom edge
                        else:
                            # Detect up boundary, Example 3064, 4759
                            result_tool_from_UD = detect_tool_from_UD(contours_number, num_of_contours, i,
                                                                      left_boundary_x, right_boundary_x, defined_height)
                            if result_tool_from_UD[0]:
                                points_touch_edge += result_tool_from_UD[0]
                                new_contours.append(result_tool_from_UD[1])

            tool_area = contour_points_in_edges(frame, points_touch_another_edge, points_touch_edge, area, mass_centres,
                                                new_contours)
            if tool_area[1]:
                leftTool_sets.append(tool_area[1])
            if tool_area[0]:
                twoConnectedTool_area_sets.append(tool_area[0])

        elif mass_centres_x > center_line_x:
            number = len(contours_number[num_of_contours])
            for i in range(number):
                # Detect whether two tools touch together
                if contours_number[num_of_contours][i][0][0] < left_boundary_x:
                    Tools_contact = detect_tools_contact(contours_number, num_of_contours, i, center_of_circle,
                                                         inner_radius, outer_radius)
                    if Tools_contact:
                        points_touch_another_edge += Tools_contact
                        cv2.circle(frame, (contours_number[num_of_contours][i][0][0],
                                           contours_number[num_of_contours][i][0][1]), radius=1,
                                   color=(130, 100, 0), thickness=4)
                # Detect if the tool comes from the right side
                else:
                    if contours_number[num_of_contours][i][0][0] > center_line_x:
                        if contours_number[num_of_contours][i][0][0] > right_boundary_x:
                            # Blue points show tool boundary between right_boundary_line_x and left_curve_path
                            cv2.circle(frame, (contours_number[num_of_contours][i][0][0],
                                               contours_number[num_of_contours][i][0][1]),
                                       radius=1, color=(200, 210, 0), thickness=4)
                            result_tool_from_LR = detect_tool_from_LR(contours_number, num_of_contours, i,
                                                                      center_of_circle, inner_radius, outer_radius)
                            if result_tool_from_LR[0]:
                                points_touch_edge += result_tool_from_LR[0]
                                new_contours.append(result_tool_from_LR[1])

                        # Try to find tools insert from up edge or bottom edge
                        else:
                            # Detect up boundary, Example 3064, 4759
                            result_tool_from_UD = detect_tool_from_UD(contours_number, num_of_contours, i,
                                                                      left_boundary_x, right_boundary_x, defined_height)
                            if result_tool_from_UD[0]:
                                points_touch_edge += result_tool_from_UD[0]
                                new_contours.append(result_tool_from_UD[1])

            tool_area = contour_points_in_edges(frame, points_touch_another_edge, points_touch_edge, area, mass_centres,
                                                new_contours)
            if tool_area[1]:
                rightTool_sets.append(tool_area[1])
            if tool_area[0]:
                twoConnectedTool_area_sets.append(tool_area[0])

    return twoConnectedTool_area_sets, leftTool_sets, rightTool_sets


'''
def process_area_output(area_list, changerate_list, frame_hasTool):
    # Later could improve this by using peak detection algorithm and sliding window


    # If during 10 frames, the view always has a tool or no tool, then we don't modify the data
    if min(area_list) > 0:
        return area_list, changerate_list

    elif max(area_list) == 0:
        return area_list, changerate_list

    else:
        for i in range(len(area_list)):
            if area_list[i] != 0:
'''

if __name__ == "__main__":
    # --------------Create Excel table--------------------------------------
    tablepath = 'Clip16_rightToolNoiseTest.xlsx'
    RowNumber = 3
    ColumnNumber = 1
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=ColumnNumber, value='Img')
    sheet.cell(row=1, column=ColumnNumber + 1, value='LightToolTot')
    sheet.cell(row=1, column=ColumnNumber + 2, value='LightTool_AreaChange')
    # --------------Create Excel table--------------------------------------

    # -------------------------------Define boundary zones--------------------------------------
    circle_center = (506, 279)
    laparoscopic_view_radius = 418
    red_color = (0, 0, 255)

    # Extra height is used to narrow the
    extra_height = 5
    width_percentage = 0.04
    # Calculate circle radii
    radius_inner = laparoscopic_view_radius - math.floor(laparoscopic_view_radius * width_percentage)
    radius_outer = laparoscopic_view_radius + 10
    radii = [radius_inner, radius_outer]
    # Calculate left and right boundaries
    box_height = math.floor(laparoscopic_view_radius * width_percentage)
    left_boundary = circle_center[0] - (laparoscopic_view_radius ** 2 - 270 ** 2) ** 0.5 + extra_height
    right_boundary = circle_center[0] + (laparoscopic_view_radius ** 2 - 270 ** 2) ** 0.5 - extra_height
    box_zone = [round(left_boundary), round(right_boundary), box_height]
    # -------------------------------Finish boundary define--------------------------------------

    leftToolTot_Previous = 0
    rightToolTot_Previous = 0
    Area_notchange = 0
    frozen_frames = 0

    # Parameters for output process
    toolinsert = 0
    counting_frame = 1
    saved_row_for_insert = 0

    for frames in range(1, 6676, 1):
        mask_frame = cv2.imread('Clip16_1M/clip16' + '_' + str(frames) + 'M.jpg')
        frame1 = cv2.imread('Clip16/clip16' + '_' + str(frames) + '.jpg')
        mask_frame = cv2.cvtColor(mask_frame, cv2.COLOR_BGR2GRAY)
        ret, contours_frame = cv2.threshold(mask_frame, 127, 255, cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(contours_frame, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        draw_defined_boundaries(frame1, circle_center, radii, box_zone)
        twoConnectedTool_area, leftTool, rightTool = identify_mask_position(frame1, contours, circle_center,
                                                                            radii, box_zone)

        if len(twoConnectedTool_area) == 1:
            twoConnectedTool_area = twoConnectedTool_area[0]
        else:
            twoConnectedTool_area = twoConnectedTool_area[1]

        leftToolTot, leftchangerate, leftTool1, leftTool2 = tool_or_noise_l(leftTool, leftToolTot_Previous)
        rightToolTot, rightchangerate, frozen_frames, rightTool1, rightTool2 = tool_or_noise_r(rightTool,
                                                                                               rightToolTot_Previous,
                                                                                               frozen_frames,
                                                                                               twoConnectedTool_area,
                                                                                               laparoscopic_view_radius)

        leftToolTot_Previous = leftToolTot
        rightToolTot_Previous = rightToolTot

        # Process output data to eliminate noise caused by sudden area change disturbance
        if toolinsert and rightToolTot:
            counting_frame += 1
        elif rightchangerate == 500:
            rightchangerate = 0
            saved_row_for_insert = RowNumber
            toolinsert = 1

        if counting_frame < 25 and rightToolTot == 0:
            rightchangerate = 0
            toolinsert = 0
            counting_frame = 1

        if counting_frame > 25 and toolinsert:
            modified_rightchangerate = 500
            sheet.cell(row=saved_row_for_insert, column=ColumnNumber + 2, value=modified_rightchangerate)
            toolinsert = 0

        if counting_frame > 25 and rightToolTot == 0:
            counting_frame = 1

        sheet.cell(row=RowNumber, column=ColumnNumber, value=('clip18_' + str(frames)))
        sheet.cell(row=RowNumber, column=ColumnNumber + 1, value=(rightToolTot / 1000))
        sheet.cell(row=RowNumber, column=ColumnNumber + 2, value=rightchangerate)
        RowNumber += 1
        print('clip25_' + str(frames))

    workbook.save(tablepath)
    print('Finish')
